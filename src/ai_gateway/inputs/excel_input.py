"""Excel input helpers.

This module keeps the dependency optional. Install with `pip install -e .[excel]`.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from ai_gateway.models import AiTask, ImageInput


def load_tasks_from_excel(path: str | Path, sheet_name: str | None = None) -> list[AiTask]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to read Excel files.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    tasks: list[AiTask] = []
    for index, values in enumerate(rows[1:], start=2):
        row: dict[str, Any] = dict(zip(headers, values))
        prompt = str(row.get("prompt") or "").strip()
        if not prompt:
            continue
        images: list[ImageInput] = []
        for key, value in row.items():
            if key.startswith("image_url") and value:
                images.append(ImageInput(type="url", value=str(value).strip()))
        tasks.append(
            AiTask(
                task_id=str(row.get("task_id") or f"row-{index}"),
                batch_id=str(row.get("batch_id") or "") or None,
                gateway=str(row.get("gateway") or "") or None,
                model=str(row.get("model") or "") or None,
                prompt=prompt,
                images=images,
                attempt=int(row.get("attempt") or 1),
                max_retries=int(row.get("max_retries") or 3),
                response_template=_parse_response_template(
                    row.get("response_template_json")
                ),
                metadata={"source": "excel", "row_number": index},
            )
        )
    return tasks


def _parse_response_template(value: Any) -> dict[str, Any] | None:
    if not value:
        return None
    if isinstance(value, dict):
        return value
    return json.loads(str(value))

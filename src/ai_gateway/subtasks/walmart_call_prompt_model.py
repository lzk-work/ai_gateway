"""Call ChatGPT-compatible model for Walmart generated prompt tasks."""

from __future__ import annotations

import argparse
import json
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import requests

from ai_gateway.clients.openai_chat_client import (
    OpenAIChatClient,
    build_responses_payload,
    extract_chat_text,
    extract_responses_text,
    is_unsupported_upstream_error,
)
from ai_gateway.config.loader import load_app_config
from ai_gateway.validators.result_validator import extract_json


PRINT_LOCK = threading.Lock()
EXCEL_FORBIDDEN_CODEPOINTS = {
    *range(0x00, 0x09),
    0x0B,
    0x0C,
    *range(0x0E, 0x20),
}


@dataclass(slots=True)
class WalmartCallPromptModelConfig:
    name: str
    input_path: str
    output_path: str
    gateways_path: str
    models_path: str
    gateway: str | None = None
    model: str | None = None
    model_candidates: list[str] = field(default_factory=list)
    max_records: int | None = None
    concurrency: int = 1
    max_tokens: int | None = 2500
    temperature: float = 0.2
    image_detail: str = "auto"
    skip_precheck_failed: bool = True
    skip_success: bool = True
    prompt_override: str | None = None
    max_retries: int = 2
    retry_delay_seconds: int = 30
    stream: bool = False
    full_outputs_dir: str | None = None
    source_excel_path: str | None = None
    excel_result_path: str | None = None
    refresh_models_on_error: bool = True


@dataclass(slots=True)
class ModelCallRecord:
    task_id: str
    batch_id: str
    sku: str | None
    status: str
    model: str
    gateway: str
    request_id: str | None
    result_text: str
    latency_ms: int | None
    attempt: int
    retryable: bool
    error_code: str | None
    error_message: str | None
    source_task: dict[str, Any]
    raw_response: dict[str, Any] | None
    created_at: str
    json_parseable: bool = False
    image_plan_count: int = 0
    validation_status: str = "not_checked"
    full_output_path: str | None = None
    row_number: int | None = None


def find_project_root(path: Path) -> Path:
    for candidate in [path.parent, *path.parents]:
        if (candidate / "src" / "ai_gateway").exists() and (candidate / "configs").exists():
            return candidate
    raise RuntimeError(f"Cannot find project root from config path: {path}")


def load_config(path: str | Path) -> WalmartCallPromptModelConfig:
    path = Path(path)
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if "input" in data or "output" in data or "execution" in data:
        project_root = find_project_root(path.resolve())
        execution = data.get("execution", {})
        model = execution.get("model", {})
        gateway = execution.get("gateway", {})
        retry = data.get("retry", {})
        data = {
            "name": data["name"],
            "input_path": data["input"]["prompt_tasks_path"],
            "output_path": data["output"]["model_results_path"],
            "gateways_path": str(project_root / "configs" / "gateways.yaml"),
            "models_path": str(project_root / "configs" / "models.yaml"),
            "gateway": gateway.get("name"),
            "model": model.get("name"),
            "model_candidates": model.get("candidates", []),
            "max_records": None,
            "max_tokens": model.get("max_tokens", 2500),
            "temperature": model.get("temperature", 0.2),
            "image_detail": model.get("image_detail", "auto"),
            "skip_precheck_failed": data.get("skip_precheck_failed", True),
            "skip_success": data.get("resume", {}).get("skip_success", True),
            "prompt_override": data.get("prompt_override"),
            "max_retries": retry.get("max_retries", 2),
            "retry_delay_seconds": retry.get("retry_delay_seconds", 30),
            "stream": bool(model.get("stream", False)),
            "full_outputs_dir": data.get("output", {}).get("full_outputs_dir"),
            "source_excel_path": data.get("input", {}).get("source_excel_path"),
            "excel_result_path": data.get("output", {}).get("excel_result_path"),
            "refresh_models_on_error": bool(model.get("refresh_on_error", True)),
        }
    return WalmartCallPromptModelConfig(**data)


class RuntimeModelPool:
    def __init__(
        self,
        gateway_config,
        preferred_model: str,
        candidates: list[str],
        enabled: bool = True,
    ) -> None:
        self.gateway_config = gateway_config
        self.enabled = enabled
        ordered = [preferred_model, *candidates]
        self.allowed = list(dict.fromkeys(item for item in ordered if item))
        self.available = list(self.allowed)
        self.current = preferred_model
        self.lock = threading.Lock()

    def current_model(self) -> str:
        with self.lock:
            return self.current

    def set_available(self, available_models: list[str]) -> None:
        with self.lock:
            filtered = [item for item in self.allowed if item in available_models]
            self.available = filtered or list(self.allowed)
            if self.current not in self.available:
                self.current = self.available[0]

    def handle_model_unavailable(self, failed_model: str) -> str | None:
        with self.lock:
            if failed_model in self.available and len(self.available) > 1:
                self.available = [item for item in self.available if item != failed_model]
                self.current = self.available[0]
                return self.current
        if not self.enabled:
            return None
        refreshed = fetch_gateway_models(self.gateway_config)
        with self.lock:
            filtered = [item for item in self.allowed if item in refreshed and item != failed_model]
            if not filtered:
                return None
            self.available = filtered
            self.current = filtered[0]
            return self.current


def fetch_gateway_models(gateway_config) -> list[str]:
    url = gateway_config.base_url.rstrip("/") + "/v1/models"
    header_name = gateway_config.auth_header or "Authorization"
    api_key = gateway_config.api_key()
    headers = {"Accept": "application/json", "User-Agent": "ai_gateway_model_probe/1.0"}
    if header_name.lower() == "authorization":
        headers[header_name] = f"Bearer {api_key}"
    else:
        headers[header_name] = api_key
    response = requests.get(url, headers=headers, timeout=60)
    response.encoding = "utf-8"
    if response.status_code >= 400:
        raise RuntimeError(f"HTTP {response.status_code}: {response.text[:1000]}")
    payload = response.json()
    return [str(item.get("id")) for item in payload.get("data", []) if isinstance(item, dict) and item.get("id")]


def run(config: WalmartCallPromptModelConfig) -> list[ModelCallRecord]:
    app_config = load_app_config(config.gateways_path, config.models_path)
    model_name = config.model or app_config.default_model
    if not model_name:
        raise RuntimeError("Missing model in config and models.yaml default_model.")
    model_config = app_config.models.get(model_name, {})
    gateway_name = config.gateway or model_config.get("gateway") or app_config.default_gateway
    gateway_config = app_config.gateways[gateway_name]
    client = OpenAIChatClient(gateway_config)
    model_pool = RuntimeModelPool(
        gateway_config,
        model_name,
        config.model_candidates,
        enabled=config.refresh_models_on_error,
    )

    source_tasks = read_jsonl(config.input_path)
    existing_rows = read_jsonl_if_exists(config.output_path)
    completed_task_ids = completed_ids(existing_rows) if config.skip_success else set()
    completed_skus = completed_sku_ids(existing_rows) if config.skip_success else set()
    pending_tasks = [
        source_task
        for source_task in source_tasks
        if source_task_id(source_task) not in completed_task_ids
        and source_task_sku(source_task) not in completed_skus
    ]
    skipped_count = len(source_tasks) - len(pending_tasks)
    if config.max_records and config.max_records > 0:
        pending_tasks = pending_tasks[: config.max_records]

    print_section("BUZZ 文本模型阶段")
    safe_print(f"任务总数: {len(source_tasks)} | 已成功跳过: {skipped_count} | 本次待处理: {len(pending_tasks)}")

    records = call_pending_tasks(pending_tasks, config, client, gateway_name, model_pool)

    merged_rows = merge_result_rows(existing_rows, records, source_tasks)
    write_jsonl_rows(merged_rows, config.output_path)
    safe_print(f"结果日志: {config.output_path}")
    write_excel_results(records_from_rows(merged_rows), config)
    if config.excel_result_path:
        safe_print(f"结果Excel: {config.excel_result_path}")
    return records


def call_pending_tasks(
    pending_tasks: list[dict[str, Any]],
    config: WalmartCallPromptModelConfig,
    client: OpenAIChatClient,
    gateway_name: str,
    model_pool: RuntimeModelPool,
) -> list[ModelCallRecord]:
    total_pending = len(pending_tasks)
    if total_pending == 0:
        return []
    concurrency = max(int(config.concurrency or 1), 1)
    concurrency = min(concurrency, total_pending)
    safe_print(f"调用并发: {concurrency}")
    if concurrency == 1:
        return call_pending_tasks_sequential(pending_tasks, config, client, gateway_name, model_pool)
    return call_pending_tasks_parallel(pending_tasks, config, client, gateway_name, model_pool, concurrency)


def call_pending_tasks_sequential(
    pending_tasks: list[dict[str, Any]],
    config: WalmartCallPromptModelConfig,
    client: OpenAIChatClient,
    gateway_name: str,
    model_pool: RuntimeModelPool,
) -> list[ModelCallRecord]:
    records: list[ModelCallRecord] = []
    total_pending = len(pending_tasks)
    for index, source_task in enumerate(pending_tasks, start=1):
        record = call_one(index, total_pending, source_task, config, client, gateway_name, model_pool)
        records.append(record)
        print_call_done(index, total_pending, record)
    return records


def call_pending_tasks_parallel(
    pending_tasks: list[dict[str, Any]],
    config: WalmartCallPromptModelConfig,
    client: OpenAIChatClient,
    gateway_name: str,
    model_pool: RuntimeModelPool,
    concurrency: int,
) -> list[ModelCallRecord]:
    total_pending = len(pending_tasks)
    indexed_tasks = list(enumerate(pending_tasks, start=1))
    results_by_index: dict[int, ModelCallRecord] = {}
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {}
        for index, source_task in indexed_tasks:
            future = executor.submit(call_one, index, total_pending, source_task, config, client, gateway_name, model_pool)
            futures[future] = index
        for future in as_completed(futures):
            index = futures[future]
            record = future.result()
            results_by_index[index] = record
            print_call_done(index, total_pending, record)
    return [results_by_index[index] for index, _ in indexed_tasks if index in results_by_index]


def print_call_start(
    index: int,
    total_pending: int,
    source_task: dict[str, Any],
    model_name: str,
    gateway_name: str,
) -> None:
    task_id = source_task_id(source_task)
    sku = source_task_sku(source_task)
    safe_print(f"[{index}/{total_pending}] 开始 | SKU={sku} | 模型={model_name} | 网关={gateway_name}")


def print_call_done(index: int, total_pending: int, record: ModelCallRecord) -> None:
    detail_parts = [
        f"[{index}/{total_pending}] 完成",
        f"SKU={record.sku}",
        f"状态={display_status(record.status)}",
    ]
    if record.latency_ms is not None:
        detail_parts.append(f"耗时={format_ms(record.latency_ms)}")
    if record.validation_status != "not_checked":
        detail_parts.append(f"校验={display_validation(record.validation_status)}")
    if record.error_message:
        detail_parts.append(f"错误={short_error(record.error_message)}")
    safe_print(" ".join(detail_parts))


def print_section(title: str) -> None:
    safe_print(f"\n=== {title} ===")


def safe_print(message: str) -> None:
    with PRINT_LOCK:
        print(message, flush=True)


def display_status(status: str) -> str:
    return {
        "success": "成功",
        "failed": "失败",
        "invalid": "校验失败",
        "submitted": "已提交",
    }.get(status, status)


def display_validation(status: str) -> str:
    return {
        "passed": "通过",
        "failed": "失败",
        "not_checked": "未校验",
    }.get(status, status)


def format_ms(value: int) -> str:
    if value >= 1000:
        return f"{value / 1000:.1f}s"
    return f"{value}ms"


def short_error(message: str, limit: int = 120) -> str:
    text = " ".join(str(message).split())
    if "<!DOCTYPE html>" in text or "<html" in text.lower():
        if "HTTP 520" in text:
            return "HTTP 520 Cloudflare HTML 错误页，上游/中转站异常"
        if "HTTP 504" in text:
            return "HTTP 504 网关超时"
        return "HTTP HTML 错误页，上游/中转站异常"
    if "No available channel for model" in text:
        return "当前模型无可用通道"
    if "model_not_found" in text:
        return "当前模型不可用或账号组不支持"
    return text if len(text) <= limit else text[:limit] + "..."
def source_task_id(source_task: dict[str, Any]) -> str:
    next_payload = source_task.get("next_task_payload") or {}
    return str(next_payload.get("task_id") or source_task.get("task_id") or "")



def source_task_sku(source_task: dict[str, Any]) -> str:
    next_payload = source_task.get("next_task_payload") or {}
    metadata = next_payload.get("metadata") or {}
    return str(metadata.get("sku") or source_task.get("sku") or "")

def source_task_row_number(source_task: dict[str, Any]) -> int | None:
    row_number = ((source_task.get("source_payload") or {}).get("row_number")) or source_task.get("row_number")
    return row_number if isinstance(row_number, int) else None


def read_text_if_exists(path: str | Path | None) -> str:
    if not path:
        return ""
    path = Path(path)
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8")
def completed_ids(rows: list[dict[str, Any]]) -> set[str]:
    ids: set[str] = set()
    for row in rows:
        if not is_completed_row(row):
            continue
        task_id = row.get("task_id")
        if task_id:
            ids.add(str(task_id))
    return ids


def completed_sku_ids(rows: list[dict[str, Any]]) -> set[str]:
    skus: set[str] = set()
    for row in rows:
        if not is_completed_row(row):
            continue
        sku = row.get("sku")
        if sku:
            skus.add(str(sku))
    return skus


def is_completed_row(row: dict[str, Any]) -> bool:
    full_output_path = row.get("full_output_path")
    if full_output_path:
        path = Path(full_output_path)
        if path.exists():
            _, _, validation_error = inspect_result_text(path.read_text(encoding="utf-8", errors="replace"))
            if validation_error is None:
                return True
            return False
    if row.get("status") != "success":
        return False
    if row.get("validation_status") != "passed" and not (
        row.get("json_parseable") is True and row.get("image_plan_count") == 6
    ):
        return False
    return True


def merge_result_rows(
    existing_rows: list[dict[str, Any]],
    new_records: list[ModelCallRecord],
    source_tasks: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    extra_rows: list[dict[str, Any]] = []
    for row in existing_rows:
        task_id = str(row.get("task_id") or "")
        if task_id:
            merged[task_id] = normalize_result_row(row)
        else:
            extra_rows.append(normalize_result_row(row))
    for record in new_records:
        merged[record.task_id] = record_to_log_row(record)

    ordered: list[dict[str, Any]] = []
    seen: set[str] = set()
    for source_task in source_tasks:
        task_id = source_task_id(source_task)
        if task_id and task_id in merged:
            ordered.append(merged[task_id])
            seen.add(task_id)
    for task_id, row in merged.items():
        if task_id not in seen:
            ordered.append(row)
    ordered.extend(extra_rows)
    return ordered


def record_to_log_row(record: ModelCallRecord) -> dict[str, Any]:
    return {
        "task_id": record.task_id,
        "batch_id": record.batch_id,
        "sku": record.sku,
        "row_number": record.row_number,
        "status": record.status,
        "model": record.model,
        "gateway": record.gateway,
        "request_id": record.request_id,
        "latency_ms": record.latency_ms,
        "attempt": record.attempt,
        "retryable": record.retryable,
        "error_code": record.error_code,
        "error_message": record.error_message,
        "created_at": record.created_at,
        "json_parseable": record.json_parseable,
        "image_plan_count": record.image_plan_count,
        "validation_status": record.validation_status,
        "full_output_path": record.full_output_path,
    }


def normalize_result_row(row: dict[str, Any]) -> dict[str, Any]:
    if "row_number" not in row:
        row_number = ((row.get("source_task") or {}).get("source_payload") or {}).get("row_number")
        if row_number is None:
            row_number = (row.get("source_task") or {}).get("row_number")
    else:
        row_number = row.get("row_number")
    return {
        "task_id": row.get("task_id"),
        "batch_id": row.get("batch_id"),
        "sku": row.get("sku"),
        "row_number": row_number,
        "status": row.get("status"),
        "model": row.get("model"),
        "gateway": row.get("gateway"),
        "request_id": row.get("request_id"),
        "latency_ms": row.get("latency_ms"),
        "attempt": row.get("attempt"),
        "retryable": row.get("retryable"),
        "error_code": row.get("error_code"),
        "error_message": row.get("error_message"),
        "created_at": row.get("created_at"),
        "json_parseable": row.get("json_parseable", False),
        "image_plan_count": row.get("image_plan_count", 0),
        "validation_status": row.get("validation_status", "not_checked"),
        "full_output_path": row.get("full_output_path"),
    }


def records_from_rows(rows: list[dict[str, Any]]) -> list[ModelCallRecord]:
    records: list[ModelCallRecord] = []
    for row in rows:
        normalized = normalize_result_row(row)
        result_text = row.get("result_text") or read_text_if_exists(normalized.get("full_output_path"))
        source_task = row.get("source_task") or {"source_payload": {"row_number": normalized.get("row_number")}}
        raw_response = row.get("raw_response")
        try:
            records.append(
                ModelCallRecord(
                    result_text=result_text or "",
                    source_task=source_task,
                    raw_response=raw_response,
                    **normalized,
                )
            )
        except TypeError:
            continue
    return records

def _call_model(
    client: OpenAIChatClient,
    payload: dict[str, Any],
    config: WalmartCallPromptModelConfig,
) -> tuple[dict[str, Any], int, str]:
    """Send one model request.

    Primary path is /v1/chat/completions (works for Gemini/Claude upstreams,
    including the gpt-5.4 alias). If BUZZ rejects the model with
    `unsupported_upstream` (OpenAI/Codex models like gpt-5.6-luna are not served
    there), automatically retry on /v1/responses with a converted payload. The
    Responses API keeps the model's reasoning trace in a separate `reasoning`
    item, so the final answer arrives clean — no "thinking instead of JSON".
    """
    try:
        if config.stream:
            response_payload, latency_ms, result_text = stream_chat_completions(client, payload)
        else:
            response_payload, latency_ms = client.chat_completions(payload)
            result_text = extract_chat_text(response_payload)
        return response_payload, latency_ms, result_text
    except RuntimeError as exc:
        if is_unsupported_upstream_error(str(exc)):
            responses_payload = build_responses_payload(payload)
            response_payload, latency_ms = client.responses_completions(responses_payload)
            result_text = extract_responses_text(response_payload)
            return response_payload, latency_ms, result_text
        raise


def call_one(
    index: int,
    total_pending: int,
    source_task: dict[str, Any],
    config: WalmartCallPromptModelConfig,
    client: OpenAIChatClient,
    gateway_name: str,
    model_pool: RuntimeModelPool,
) -> ModelCallRecord:
    next_payload = source_task.get("next_task_payload") or {}
    task_id = str(next_payload.get("task_id") or source_task.get("task_id") or "")
    batch_id = str(next_payload.get("batch_id") or source_task.get("batch_id") or "")
    metadata = next_payload.get("metadata") or {}
    sku = metadata.get("sku") or source_task.get("sku")
    print_call_start(index, total_pending, source_task, model_pool.current_model(), gateway_name)

    if config.skip_precheck_failed and next_payload.get("precheck_status") == "failed":
        model_name = model_pool.current_model()
        return build_error_record(
            source_task,
            task_id,
            batch_id,
            sku,
            gateway_name,
            model_name,
            "PrecheckFailed",
            json.dumps(next_payload.get("precheck_errors", []), ensure_ascii=False),
        )

    last_error: Exception | None = None
    latency_ms: int | None = None
    model_name = model_pool.current_model()
    last_attempt = 1
    max_attempts = max(config.max_retries + 1, 1)
    for attempt in range(1, max_attempts + 1):
        last_attempt = attempt
        model_name = model_pool.current_model()
        payload = build_chat_payload(
            next_payload=next_payload,
            source_task=source_task,
            prompt_override=config.prompt_override,
            model_name=model_name,
            max_tokens=config.max_tokens,
            temperature=config.temperature,
            image_detail=config.image_detail,
        )
        try:
            response_payload, latency_ms, result_text = _call_model(client, payload, config)
            full_output_path = save_full_output(config, task_id, sku, model_name, result_text)
            json_parseable, image_plan_count, validation_error = inspect_result_text(result_text)
            validation_status = "passed" if not validation_error else "failed"
            status = "success" if validation_status == "passed" else "invalid"
            return ModelCallRecord(
                task_id=task_id,
                batch_id=batch_id,
                sku=sku,
                status=status,
                model=model_name,
                gateway=gateway_name,
                request_id=response_payload.get("id"),
                result_text=result_text,
                latency_ms=latency_ms,
                attempt=attempt,
                retryable=False,
                json_parseable=json_parseable,
                image_plan_count=image_plan_count,
                validation_status=validation_status,
                full_output_path=full_output_path,
                error_code=None,
                error_message=validation_error,
                source_task=source_task,
                raw_response=response_payload,
                created_at=datetime.now().isoformat(timespec="seconds"),
                row_number=source_task_row_number(source_task),
            )
        except Exception as exc:
            last_error = exc
            if is_model_unavailable_error_message(str(exc)):
                next_model = model_pool.handle_model_unavailable(model_name)
                if next_model and attempt < max_attempts:
                    safe_print(f"模型切换: {model_name} -> {next_model} | 原因=模型通道不可用")
                    continue
            if attempt >= max_attempts or not is_retryable_error_message(str(exc)):
                break
            time.sleep(config.retry_delay_seconds)

    error = last_error or RuntimeError("unknown model call error")
    return build_error_record(
        source_task,
        task_id,
        batch_id,
        sku,
        gateway_name,
        model_name,
        error.__class__.__name__,
        str(error),
        attempt=last_attempt,
        latency_ms=latency_ms,
    )


def build_chat_payload(
    next_payload: dict[str, Any],
    source_task: dict[str, Any],
    prompt_override: str | None,
    model_name: str,
    max_tokens: int | None,
    temperature: float,
    image_detail: str,
) -> dict[str, Any]:
    prompt = prompt_override or str(next_payload.get("prompt") or "")
    if prompt_override:
        prompt = render_prompt_override(prompt_override, source_task)
    content: list[dict[str, Any]] = [
        {"type": "text", "text": prompt}
    ]
    for image in next_payload.get("images", []):
        if not isinstance(image, dict) or image.get("type") != "url":
            continue
        image_url: dict[str, Any] = {"url": image.get("value")}
        if image_detail and image_detail != "auto":
            image_url["detail"] = image_detail
        content.append({"type": "image_url", "image_url": image_url})

    payload: dict[str, Any] = {
        "model": model_name,
        "messages": [{"role": "user", "content": content}],
        "temperature": temperature,
    }
    if max_tokens:
        payload["max_tokens"] = max_tokens
    return payload


CONTINUE_PROMPT = (
    "继续。请直接以字符 { 开头输出完整的 JSON 对象，"
    "不要任何解释、开场白或计划性文字。image_plan 必须包含 6 个对象。"
)


def send_request(client, payload, stream):
    """Send a chat completion request (stream or not) and return (response, latency_ms, text)."""
    if stream:
        return stream_chat_completions(client, payload)
    response_payload, latency_ms = client.chat_completions(payload)
    return response_payload, latency_ms, extract_chat_text(response_payload)


def is_pure_thinking(text: str) -> bool:
    """Heuristic: model returned only a thinking/plan preamble without any JSON.

    Used to decide whether a 'continue' follow-up turn can salvage the call.
    We only treat short, JSON-less responses as pure thinking; truncated JSON
    (long or containing '{') is left as a normal invalid result.
    """
    parsed, err = extract_json(text)
    if err is None:
        return False
    stripped = text.strip()
    if len(stripped) > 2000:
        return False
    if "{" in stripped:
        return False
    return True


def build_continue_payload(
    next_payload,
    source_task,
    prompt_override,
    model_name,
    max_tokens,
    temperature,
    image_detail,
    previous_text,
):
    """Build a multi-turn payload that continues a thinking-only first turn.

    messages = [original user (with images), assistant(thinking), user(continue)]
    """
    prompt = prompt_override or str(next_payload.get("prompt") or "")
    if prompt_override:
        prompt = render_prompt_override(prompt_override, source_task)
    content = [{"type": "text", "text": prompt}]
    for image in next_payload.get("images", []):
        if not isinstance(image, dict) or image.get("type") != "url":
            continue
        image_url = {"url": image.get("value")}
        if image_detail and image_detail != "auto":
            image_url["detail"] = image_detail
        content.append({"type": "image_url", "image_url": image_url})
    messages = [
        {"role": "user", "content": content},
        {"role": "assistant", "content": previous_text},
        {"role": "user", "content": CONTINUE_PROMPT},
    ]
    payload = {"model": model_name, "messages": messages, "temperature": temperature}
    if max_tokens:
        payload["max_tokens"] = max_tokens
    return payload


def render_prompt_override(template: str, source_task: dict[str, Any]) -> str:
    row = ((source_task.get("source_payload") or {}).get("row") or {})
    values = {
        "sku": source_task.get("sku") or row.get("开发SKU") or "",
        "title": row.get("标题") or (source_task.get("source_payload") or {}).get("title") or "",
        "bullets": row.get("五点") or (source_task.get("source_payload") or {}).get("bullets") or "",
        "category": row.get("类目") or "",
    }
    rendered = template
    for key, value in values.items():
        rendered = rendered.replace("{{" + key + "}}", str(value or ""))
    return rendered


def build_error_record(
    source_task: dict[str, Any],
    task_id: str,
    batch_id: str,
    sku: str | None,
    gateway_name: str,
    model_name: str,
    error_code: str,
    error_message: str,
    attempt: int = 1,
    latency_ms: int | None = None,
) -> ModelCallRecord:
    return ModelCallRecord(
        task_id=task_id,
        batch_id=batch_id,
        sku=sku,
        status="failed",
        model=model_name,
        gateway=gateway_name,
        request_id=None,
        result_text="",
        latency_ms=latency_ms,
        attempt=attempt,
        retryable=is_retryable_error_message(error_message),
        json_parseable=False,
        image_plan_count=0,
        validation_status="not_checked",
        full_output_path=None,
        error_code=error_code,
        error_message=error_message,
        source_task=source_task,
        raw_response=None,
        created_at=datetime.now().isoformat(timespec="seconds"),
        row_number=source_task_row_number(source_task),
    )


def is_retryable_error_message(message: str) -> bool:
    retryable_markers = ("HTTP 429", "HTTP 500", "HTTP 502", "HTTP 503", "HTTP 504", "HTTP 525", "timeout", "timed out")
    non_retryable_markers = ("HTTP 400", "HTTP 401", "HTTP 403", "model_not_found", "invalid_request")
    if any(marker in message for marker in non_retryable_markers):
        return False
    return any(marker in message for marker in retryable_markers)


def is_model_unavailable_error_message(message: str) -> bool:
    markers = (
        "model_not_found",
        "No available channel for model",
        "is not supported by any configured account",
    )
    return any(marker in message for marker in markers)


def stream_chat_completions(
    client: OpenAIChatClient,
    payload: dict[str, Any],
) -> tuple[dict[str, Any], int]:
    payload = dict(payload)
    payload["stream"] = True
    url = client.gateway.base_url.rstrip("/") + "/v1/chat/completions"
    started = time.perf_counter()
    text_parts: list[str] = []
    response_id: str | None = None
    model_name: str | None = None
    with requests.post(
        url,
        headers={
            "Authorization": f"Bearer {client.gateway.api_key()}",
            "Content-Type": "application/json",
            "Accept": "text/event-stream",
            "User-Agent": "Mozilla/5.0",
        },
        json=payload,
        stream=True,
        timeout=client.gateway.timeout_seconds,
    ) as response:
        latency_ms = int((time.perf_counter() - started) * 1000)
        response.encoding = "utf-8"
        if response.status_code >= 400:
            raise RuntimeError(f"HTTP {response.status_code}: {response.text[:1000]}")
        for raw_line in response.iter_lines(decode_unicode=True):
            if not raw_line:
                continue
            line = raw_line.strip()
            if not line.startswith("data:"):
                continue
            data = line.removeprefix("data:").strip()
            if data == "[DONE]":
                break
            try:
                event = json.loads(data)
            except json.JSONDecodeError:
                continue
            response_id = response_id or event.get("id")
            model_name = model_name or event.get("model")
            text_parts.append(extract_stream_delta(event))
    return (
        {
            "id": response_id,
            "model": model_name,
            "choices": [{"message": {"content": "".join(text_parts)}}],
        },
        latency_ms,
        "".join(text_parts),
    )


def extract_stream_delta(event: dict[str, Any]) -> str:
    choices = event.get("choices") or []
    if not choices:
        return ""
    delta = choices[0].get("delta") or {}
    content = delta.get("content")
    if isinstance(content, str):
        return content
    return ""


def inspect_result_text(text: str) -> tuple[bool, int, str | None]:
    raw_mojibake_error = mojibake_reason(text)
    if raw_mojibake_error:
        return False, 0, f"{raw_mojibake_error} in full response"
    parsed, parse_error = extract_json(text)
    if parse_error or not isinstance(parsed, dict):
        return False, 0, parse_error or "response JSON is not an object"
    image_plan = parsed.get("image_plan")
    image_plan_count = len(image_plan) if isinstance(image_plan, list) else 0
    if image_plan_count != 6:
        return True, image_plan_count, f"image_plan count expected 6, got {image_plan_count}"
    mojibake_error = detect_mojibake(parsed)
    if mojibake_error:
        return True, image_plan_count, mojibake_error
    return True, image_plan_count, None


def detect_mojibake(value: Any) -> str | None:
    for path, text in iter_strings(value):
        reason = mojibake_reason(text)
        if reason:
            return f"{reason} at {path}"
    return None


def iter_strings(value: Any, path: str = "$"):
    if isinstance(value, dict):
        for key, child in value.items():
            yield from iter_strings(child, f"{path}.{key}")
    elif isinstance(value, list):
        for index, child in enumerate(value):
            yield from iter_strings(child, f"{path}[{index}]")
    elif isinstance(value, str):
        yield path, value


def mojibake_reason(text: str) -> str | None:
    if not text:
        return None
    if "\ufffd" in text or "锟斤拷" in text:
        return "疑似乱码: replacement characters"
    c1_controls = sum(1 for char in text if 0x80 <= ord(char) <= 0x9F)
    if c1_controls >= 10 or (c1_controls >= 5 and c1_controls / len(text) >= 0.01):
        return "疑似乱码: C1 control characters"
    if len(text) < 20:
        return None
    mojibake_chars = sum(1 for char in text if char in "ÂÃÄÅÆÇÈÉÊËÌÍÎÏÐÑÒÓÔÕÖØÙÚÛÜÝÞßàáâãäåæçèéêëìíîïðñòóôõöøùúûüýþÿ")
    cjk_chars = sum(1 for char in text if "\u4e00" <= char <= "\u9fff")
    if cjk_chars == 0 and mojibake_chars >= 8 and mojibake_chars / len(text) >= 0.08:
        return "疑似乱码: UTF-8 text decoded as Latin-1"
    return None


def save_full_output(
    config: WalmartCallPromptModelConfig,
    task_id: str,
    sku: str | None,
    model_name: str,
    result_text: str,
) -> str | None:
    if not config.full_outputs_dir or not result_text:
        return None
    business_id = sku or task_id
    safe_business_id = "".join(
        ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in business_id
    )
    path = Path(config.full_outputs_dir) / f"{safe_business_id}__{model_name}.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(result_text, encoding="utf-8")
    return str(path)


def write_excel_results(
    records: list[ModelCallRecord],
    config: WalmartCallPromptModelConfig,
) -> None:
    if not config.source_excel_path or not config.excel_result_path:
        return
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to write Excel result files.") from exc

    source_path = Path(config.source_excel_path)
    output_path = Path(config.excel_result_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source_path)
    sheet = workbook.active
    header_row = 1
    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[header_row]
        if cell.value is not None and str(cell.value).strip()
    }
    result_columns = [
        "ai_status",
        "ai_validation_status",
        "ai_result_json",
        "ai_error_message",
        "ai_processed_at",
    ]
    for column_name in result_columns:
        if column_name not in headers:
            headers[column_name] = sheet.max_column + 1
            sheet.cell(row=header_row, column=headers[column_name], value=column_name)

    for record in records:
        row_number = record.row_number or ((record.source_task.get("source_payload") or {}).get("row_number")) or record.source_task.get("row_number")
        if not isinstance(row_number, int):
            continue
        values = {
            "ai_status": record.status,
            "ai_validation_status": record.validation_status,
            "ai_result_json": build_excel_result_value(record.result_text),
            "ai_error_message": record.error_message,
            "ai_processed_at": record.created_at,
        }
        for column_name, value in values.items():
            sheet.cell(row=row_number, column=headers[column_name], value=sanitize_excel_value(value))

    workbook.save(output_path)


def sanitize_excel_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return "".join(char for char in value if ord(char) not in EXCEL_FORBIDDEN_CODEPOINTS)



def build_excel_result_value(result_text: str) -> str:
    if not result_text:
        return ""
    parsed, parse_error = extract_json(result_text)
    if parse_error or not isinstance(parsed, dict):
        text = result_text
        return text if len(text) <= 30000 else "过长，请查看JSONL或full_outputs完整结果"

    full_text = json.dumps(parsed, ensure_ascii=False)
    if len(full_text) <= 30000:
        return full_text

    structure = strip_data_keep_structure(
        parsed,
        keep_data_keys={"image_plan", "global_prompt_restrictions"},
    )
    structure_text = json.dumps(structure, ensure_ascii=False)
    if len(structure_text) <= 30000:
        return structure_text

    return "过长，请查看JSONL或full_outputs完整结果"

def build_excel_image_plan_value(result_text: str) -> str:
    if not result_text:
        return ""
    parsed, parse_error = extract_json(result_text)
    if parse_error or not isinstance(parsed, dict):
        return "JSON解析失败"
    image_plan = parsed.get("image_plan")
    if not isinstance(image_plan, list):
        return "缺少image_plan"
    text = json.dumps(image_plan, ensure_ascii=False)
    if len(text) > 30000:
        return "过长，请查看ai_full_output_path完整JSON"
    return text


def build_excel_structure_value(result_text: str) -> str:
    if not result_text:
        return ""
    parsed, parse_error = extract_json(result_text)
    if parse_error or not isinstance(parsed, dict):
        return "JSON解析失败"
    structure = strip_data_keep_structure(parsed, keep_data_keys={"image_plan"})
    text = json.dumps(structure, ensure_ascii=False)
    if len(text) > 30000:
        return "过长，请查看ai_full_output_path完整JSON"
    return text


def strip_data_keep_structure(value: Any, keep_data_keys: set[str] | None = None, current_key: str | None = None) -> Any:
    keep_data_keys = keep_data_keys or set()
    if current_key in keep_data_keys:
        return value
    if isinstance(value, dict):
        return {
            key: strip_data_keep_structure(child, keep_data_keys, key)
            for key, child in value.items()
        }
    if isinstance(value, list):
        return []
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None
    if isinstance(value, str):
        return ""
    return None


def build_result_preview(result_text: str) -> str:
    if not result_text:
        return ""
    parsed, parse_error = extract_json(result_text)
    if parse_error or not isinstance(parsed, dict):
        return result_text[:30000]
    preview = {
        "product_analysis": strip_data_keep_structure(parsed.get("product_analysis", {})),
        "image_plan_count": len(parsed.get("image_plan", []))
        if isinstance(parsed.get("image_plan"), list)
        else 0,
        "final_checklist": parsed.get("final_checklist", {}),
    }
    return json.dumps(preview, ensure_ascii=False)[:30000]


def read_jsonl_if_exists(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if not path.exists():
        return []
    return read_jsonl(path)

def read_jsonl(path: str | Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with Path(path).open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def write_jsonl_rows(rows: list[dict[str, Any]], path: str | Path) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")

def write_jsonl(records: list[ModelCallRecord], path: str | Path) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record_to_log_row(record), ensure_ascii=False) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Call model for Walmart prompt tasks.")
    parser.add_argument("--config", required=True, help="Path to subtask config JSON.")
    args = parser.parse_args()
    records = run(load_config(args.config))
    print(f"called={len(records)}")
    if records:
        success_count = sum(1 for item in records if item.status == "success")
        print(f"success={success_count}")
        print(f"failed={len(records) - success_count}")


if __name__ == "__main__":
    main()





















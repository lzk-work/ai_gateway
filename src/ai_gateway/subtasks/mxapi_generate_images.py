"""Generate images with MXAPI gpt-image-2 from prepared workbook rows."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ai_gateway.clients.mxapi_image_client import MxapiImageClient
from ai_gateway.config.loader import load_app_config
from ai_gateway.validators.result_validator import extract_json


@dataclass(slots=True)
class MxapiGenerateImagesConfig:
    name: str
    gateways_path: str
    models_path: str
    gateway: str
    submit_endpoint: str
    query_endpoint: str
    model: str
    aspect_ratio: str
    quality: str
    resolution: str
    input_excel_path: str
    input_sheet_name: str
    model_results_path: str
    output_excel_path: str
    output_results_path: str
    checkpoint_path: str
    download_dir: str
    raw_responses_dir: str
    columns: dict[str, str]
    prompt_mode: str = "buzz"
    prompt_column: str = "生成提示词"
    max_records: int | None = None
    concurrency: int = 1
    poll_interval_seconds: int = 5
    max_wait_seconds: int = 300
    submit_delay_seconds: float = 1.5
    download_timeout_seconds: int = 60
    max_submit_retries: int = 3
    max_download_retries: int = 2
    retry_delay_seconds: int = 5
    skip_success: bool = True
    poll_existing_task_id: bool = True
    image_type_order: list[str] = field(default_factory=list)
    desired_count: int | None = None
    require_success_results_path: str | None = None


@dataclass(slots=True)
class ImageGenerationRecord:
    row_number: int
    sku: str
    image_name: str
    image_type: str | None
    image_number: int | None
    status: str
    task_id: str | None
    reference_image: str | None
    generated_image_url: str | None
    downloaded_path: str | None
    file_size: int | None
    error_message: str | None
    retryable: bool
    submit_latency_ms: int | None
    poll_count: int
    total_wait_seconds: int | None
    created_at: str


def find_project_root(path: Path) -> Path:
    for candidate in [path.parent, *path.parents]:
        if (candidate / "src" / "ai_gateway").exists() and (candidate / "configs").exists():
            return candidate
    raise RuntimeError(f"Cannot find project root from config path: {path}")


def load_config(path: str | Path) -> MxapiGenerateImagesConfig:
    path = Path(path)
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    project_root = find_project_root(path.resolve())
    execution = data["execution"]
    gateway = execution["gateway"]
    model = execution["model"]
    limits = data.get("limits", {})
    retry = data.get("retry", {})
    resume = data.get("resume", {})
    config = MxapiGenerateImagesConfig(
        name=data["name"],
        gateways_path=str(project_root / "configs" / "gateways.yaml"),
        models_path=str(project_root / "configs" / "models.yaml"),
        gateway=gateway["name"],
        submit_endpoint=gateway.get("endpoint_submit", "/api/v2/gpt-image-2"),
        query_endpoint=gateway.get("endpoint_query", "/api/v2/gpt-image/task"),
        model=model.get("name", "gpt-image-2"),
        aspect_ratio=model.get("aspect_ratio", "1:1"),
        quality=model.get("quality", "low"),
        resolution=model.get("resolution", "1K"),
        input_excel_path=data["input"]["excel_path"],
        input_sheet_name=data["input"].get("sheet_name", "Sheet1"),
        model_results_path=data["input"]["model_results_path"],
        output_excel_path=data["output"]["excel_path"],
        output_results_path=data["output"]["results_path"],
        checkpoint_path=data["output"].get("checkpoint_path", data["output"]["results_path"]),
        download_dir=data["output"]["download_dir"],
        raw_responses_dir=data["output"].get("raw_responses_dir", ""),
        columns=data["columns"],
        poll_interval_seconds=int(limits.get("poll_interval_seconds", 5)),
        max_wait_seconds=int(limits.get("max_wait_seconds", 300)),
        submit_delay_seconds=float(limits.get("submit_delay_seconds", 1.5)),
        download_timeout_seconds=int(limits.get("download_timeout_seconds", 60)),
        max_submit_retries=int(retry.get("max_submit_retries", 3)),
        max_download_retries=int(retry.get("max_download_retries", 2)),
        retry_delay_seconds=int(retry.get("retry_delay_seconds", 5)),
        skip_success=bool(resume.get("skip_success", True)),
        poll_existing_task_id=bool(resume.get("poll_existing_task_id", True)),
    )
    config.prompt_mode = str(data.get("prompt_mode", "buzz")).strip().lower()
    config.prompt_column = str(data.get("prompt_column", "生成提示词")).strip()
    # 读业务级总配置（subtasks/walmart_image_prompt/config.json）的 image_selection
    image_type_order: list[str] = []
    desired_count: int | None = None
    task_root = project_root / "subtasks" / "walmart_image_prompt"
    total_cfg_path = task_root / "config.json"
    if total_cfg_path.exists():
        try:
            total_data = json.loads(total_cfg_path.read_text(encoding="utf-8-sig"))
            selection = total_data.get("image_selection") or {}
            image_type_order = [str(t).strip() for t in selection.get("image_type_order", []) if str(t).strip()]
            dc = selection.get("desired_count")
            desired_count = int(dc) if isinstance(dc, int) and dc > 0 else None
        except Exception:
            pass
    config.image_type_order = image_type_order
    config.desired_count = desired_count
    config.require_success_results_path = data.get("require_success_results_path")
    return config



class CheckpointStore:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        self.lock = threading.Lock()
        self.records: dict[str, dict[str, Any]] = {}
        for row in read_jsonl_if_exists(self.path):
            key = record_key(row)
            if key:
                self.records[key] = row

    def rows(self) -> list[dict[str, Any]]:
        with self.lock:
            return list(self.records.values())

    def upsert(self, record: ImageGenerationRecord) -> None:
        with self.lock:
            row = asdict(record)
            self.records[record_key(row)] = row
            self.append_locked(row)

    def append_locked(self, row: dict[str, Any]) -> None:
        """追加写一行（取代全量重写 + os.replace）：锁窗口从“重写整文件”缩到“追加一行”，
        缓解外部进程锁导致的 PermissionError WinError5。写失败仅告警，不崩整个进程。"""
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            with self.path.open("a", encoding="utf-8") as handle:
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
        except OSError as exc:
            print(f"[checkpoint] 写入失败(不中断): {exc}", flush=True)


def record_key(row: dict[str, Any]) -> str:
    sku = row.get("sku")
    image_name = row.get("image_name")
    return f"{sku}::{image_name}" if sku and image_name else ""


class RowProgress:
    """图片行级全局进度：跨 SKU 线程共享，线程安全地推进“已完成行/总行数”。

    并发下“开始第 N 行”没有唯一含义，故只打完成计数 + 状态，便于观察整体进度。
    """

    _STATUS_TEXT = {
        "success": "成功",
        "pending": "超时待定",
        "skipped": "跳过",
        "failed_permanent": "永久失败",
        "failed": "失败",
        "submitted": "已提交",
        "blocked": "依赖未满足",
    }

    def __init__(self, total: int) -> None:
        self.total = max(int(total), 0)
        self.done = 0
        self.lock = threading.Lock()

    def advance(self, status: str, sku: str) -> None:
        with self.lock:
            self.done += 1
            text = self._STATUS_TEXT.get(status, status)
            print(f"[进度 {self.done}/{self.total} 行] SKU={sku} | {text}", flush=True)

def run(config: MxapiGenerateImagesConfig) -> list[ImageGenerationRecord]:
    app_config = load_app_config(config.gateways_path, config.models_path)
    gateway_config = app_config.gateways[config.gateway]
    client = MxapiImageClient(gateway_config, config.submit_endpoint, config.query_endpoint)

    prompt_map = load_prompt_map(config.model_results_path)
    rows, workbook, sheet, headers = load_work_rows(config)
    checkpoint_store = CheckpointStore(config.checkpoint_path)
    existing_records = checkpoint_store.rows()
    rows = apply_checkpoint_to_rows(rows, existing_records)
    completed = completed_keys(existing_records) if config.skip_success else set()
    pending_rows = [row for row in rows if row_key(row) not in completed]
    if config.max_records and config.max_records > 0:
        pending_rows = limit_rows_by_sku(pending_rows, config.max_records)

    progress = RowProgress(len(rows))
    blocked_records: list[ImageGenerationRecord] = []
    require_path = config.require_success_results_path
    if require_path and Path(require_path).exists():
        success_skus = _load_success_skus(require_path)
        kept_rows: list[dict[str, Any]] = []
        for row in pending_rows:
            sku = str(row.get("sku") or "").strip()
            if sku in success_skus:
                kept_rows.append(row)
            else:
                # 依赖未满足（如主图未成功）：不提交 MXAPI，不消耗生成积分；
                # blocked 非终态(retryable)，续跑会重新判定，主图成功后自动补生成。
                rec = build_blocked_record(row)
                blocked_records.append(rec)
                checkpoint_store.upsert(rec)
                progress.advance(rec.status, sku)
        pending_rows = kept_rows
        print(
            f"依赖主图成功: 主图成功SKU={len(success_skus)} | "
            f"因依赖未满足跳过副图行={len(blocked_records)}",
            flush=True,
        )
    elif require_path:
        print(f"依赖主图成功: 主图结果文件不存在({require_path})，不施加依赖跳过", flush=True)

    print("\n=== MXAPI 图片生成阶段 ===", flush=True)
    print(
        f"图片任务总数: {len(rows)} 行 / {count_skus(rows)} 个 SKU | "
        f"已成功跳过: {len(rows) - len(pending_rows) - len(blocked_records)} 行 | "
        f"依赖未满足(跳过): {len(blocked_records)} 行 | "
        f"本次待处理: {len(pending_rows)} 行 / {count_skus(pending_rows)} 个 SKU "
        f"(max_records={config.max_records} 个 SKU)",
        flush=True,
    )
    records = process_rows(pending_rows, prompt_map, config, client, checkpoint_store, progress)
    records.extend(blocked_records)
    merged = merge_records(checkpoint_store.rows(), records, rows)
    write_jsonl_rows(merged, config.output_results_path)
    write_excel(workbook, sheet, headers, merged, config)
    return records


def _load_success_skus(results_path: str | Path) -> set[str]:
    """从另一结果 JSONL 中收集 status==success 的 SKU 集合，用于「依赖主图成功才生成副图」判定。"""
    skus: set[str] = set()
    for row in read_jsonl_if_exists(results_path):
        if row.get("status") == "success" and row.get("sku"):
            skus.add(str(row["sku"]).strip())
    return skus


def load_work_rows(config: MxapiGenerateImagesConfig):
    input_path = Path(config.input_excel_path)
    workbook = load_workbook(input_path)
    sheet = workbook[config.input_sheet_name]
    headers = header_map(sheet)
    required = [
        config.columns["sku"],
        config.columns["reference_image"],
        config.columns["image_name"],
        config.columns["status"],
        config.columns["task_id"],
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError("Missing image input headers: " + ", ".join(missing))

    rows: list[dict[str, Any]] = []
    image_type_header = config.columns.get("image_type")
    has_image_type = bool(image_type_header) and image_type_header in headers
    prompt_header = config.columns.get("prompt")
    has_prompt = bool(prompt_header) and prompt_header in headers
    for row_number in range(2, sheet.max_row + 1):
        sku = cell_text(sheet, row_number, headers[config.columns["sku"]])
        image_name = cell_text(sheet, row_number, headers[config.columns["image_name"]])
        reference_image = cell_text(sheet, row_number, headers[config.columns["reference_image"]])
        status = cell_text(sheet, row_number, headers[config.columns["status"]])
        task_id = cell_text(sheet, row_number, headers[config.columns["task_id"]])
        if not sku or not image_name:
            continue
        row = {
            "row_number": row_number,
            "sku": sku,
            "image_name": image_name,
            "reference_image": reference_image,
            "status": status,
            "task_id": task_id,
            "image_number": parse_image_number(image_name),
        }
        if has_image_type:
            row["image_type"] = cell_text(sheet, row_number, headers[image_type_header])
        if has_prompt:
            row["prompt"] = cell_text(sheet, row_number, headers[prompt_header])
        rows.append(row)
    return rows, workbook, sheet, headers


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def cell_text(sheet, row: int, column: int) -> str:
    value = sheet.cell(row, column).value
    return "" if value is None else str(value).strip()


def row_key(row: dict[str, Any]) -> str:
    return f"{row['sku']}::{row['image_name']}"


def count_skus(rows: list[dict[str, Any]]) -> int:
    return len({str(row.get("sku") or "").strip() for row in rows if str(row.get("sku") or "").strip()})


def limit_rows_by_sku(rows: list[dict[str, Any]], max_records: int | None) -> list[dict[str, Any]]:
    """按 SKU 为单位截断：最多保留前 max_records 个 SKU 的全部行。

    max_records 语义为“源数据行数（SKU 数）”，每个 SKU 展开的图片行作为一个整体，
    要么全部保留、要么全部截断，避免同一 SKU 只生成部分副图。
    """
    if not max_records or max_records <= 0:
        return rows
    seen: set[str] = set()
    selected: list[dict[str, Any]] = []
    for row in rows:
        sku = str(row.get("sku") or "").strip()
        if not sku:
            continue
        if sku not in seen:
            if len(seen) >= max_records:
                continue
            seen.add(sku)
        selected.append(row)
    return selected


def completed_keys(rows: list[dict[str, Any]]) -> set[str]:
    return {record_key(row) for row in rows if is_terminal_skippable(row)}


def is_terminal_skippable(row: dict[str, Any]) -> bool:
    """续跑时可直接跳过的终态：成功 或 已主动跳过（目标张数已满足）。"""
    if row.get("status") == "skipped":
        return True
    return is_completed_success(row)


def is_completed_success(row: dict[str, Any]) -> bool:
    if row.get("status") != "success":
        return False
    key = record_key(row)
    if not key:
        return False
    downloaded_path = row.get("downloaded_path")
    if not downloaded_path:
        return False
    path = Path(str(downloaded_path))
    if not path.is_file():
        return False
    expected_size = row.get("file_size")
    actual_size = path.stat().st_size
    if actual_size <= 0:
        return False
    if isinstance(expected_size, int) and expected_size > 0 and actual_size != expected_size:
        return False
    return True


def apply_checkpoint_to_rows(rows: list[dict[str, Any]], checkpoint_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    checkpoint_by_key = {record_key(row): row for row in checkpoint_rows if record_key(row)}
    enriched: list[dict[str, Any]] = []
    for row in rows:
        merged = dict(row)
        saved = checkpoint_by_key.get(row_key(row))
        if saved:
            if saved.get("task_id"):
                merged["task_id"] = str(saved["task_id"])
            if saved.get("status"):
                merged["status"] = str(saved["status"])
        enriched.append(merged)
    return enriched


def parse_image_number(image_name: str) -> int | None:
    match = re.search(r"sub(\d+)", image_name, re.IGNORECASE)
    return int(match.group(1)) if match else None


def load_prompt_map(model_results_path: str | Path) -> dict[str, dict[int, str]]:
    prompt_map: dict[str, dict[int, str]] = {}
    for row in read_jsonl_if_exists(model_results_path):
        if row.get("status") != "success" or row.get("validation_status") != "passed":
            continue
        sku = row.get("sku")
        full_output_path = row.get("full_output_path")
        if not sku or not full_output_path:
            continue
        text = Path(full_output_path).read_text(encoding="utf-8") if Path(full_output_path).exists() else ""
        parsed, parse_error = extract_json(text)
        if parse_error or not isinstance(parsed, dict):
            continue
        global_restrictions = parsed.get("global_prompt_restrictions")
        global_text = format_prompt_part(global_restrictions)
        item_map: dict[int, str] = {}
        for item in parsed.get("image_plan", []):
            if not isinstance(item, dict):
                continue
            image_number = item.get("image_number")
            if not image_number:
                continue
            prompt = (
                item.get("ai_image_generation_prompt")
                or item.get("image_generation_prompt")
                or item.get("prompt")
                or item.get("design_strategy")
                or ""
            )
            prompt = str(prompt).strip()
            if global_text:
                prompt = f"{prompt}\n\nGlobal Prompt Restrictions:\n{global_text}".strip()
            item_map[int(image_number)] = prompt
        prompt_map[str(sku).strip()] = item_map
    return prompt_map


def format_prompt_part(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, str):
        return value.strip()
    return json.dumps(value, ensure_ascii=False)


def process_rows(
    rows: list[dict[str, Any]],
    prompt_map: dict[str, dict[int, str]],
    config: MxapiGenerateImagesConfig,
    client: MxapiImageClient,
    checkpoint_store: CheckpointStore,
    progress: RowProgress,
) -> list[ImageGenerationRecord]:
    if not rows:
        return []
    # 按 SKU 分组：SKU 内按 image_type_order 串行，SKU 间并发
    by_sku: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        sku = str(row.get("sku") or "").strip()
        if sku:
            by_sku.setdefault(sku, []).append(row)
    order_index = {t: i for i, t in enumerate(config.image_type_order)}
    def sort_key(r: dict[str, Any]) -> int:
        it = str(r.get("image_type") or "").strip()
        return order_index.get(it, len(order_index))

    concurrency = max(int(config.concurrency or 1), 1)
    concurrency = min(concurrency, max(len(by_sku), 1))
    print(f"图片生成并发(SKU级): {concurrency} | SKU数={len(by_sku)}", flush=True)

    def process_sku(sku: str, sku_rows: list[dict[str, Any]]) -> list[ImageGenerationRecord]:
        sku_rows = sorted(sku_rows, key=sort_key)
        records: list[ImageGenerationRecord] = []
        # 重启续跑：该 SKU 已有成功数从 checkpoint 统计（pending_rows 只含未成功行）
        satisfied = sum(
            1
            for saved in checkpoint_store.rows()
            if str(saved.get("sku") or "").strip() == sku and saved.get("status") == "success"
        )
        print(
            f"开始处理 SKU={sku} | 已有成功={satisfied} | 目标={config.desired_count or '-'} | 候选行={len(sku_rows)}",
            flush=True,
        )
        for position, row in enumerate(sku_rows, start=1):
            if config.desired_count and satisfied >= config.desired_count:
                # 已满足目标张数：跳过该 SKU 剩余类型行（不生成、不判失败）
                rec = build_skipped_record(row)
                records.append(rec)
                progress.advance(rec.status, sku)
                continue
            if row.get("status") == "failed_permanent":
                # 跨重启：该类型已确认永久失败，不再重新提交；后续行（后位类型）继续补位
                rec = build_permanent_record(row, "已被安全策略拦截，永久放弃")
                records.append(rec)
                progress.advance(rec.status, sku)
                continue
            rec = process_one(position, len(sku_rows), row, prompt_map, config, client, checkpoint_store)
            records.append(rec)
            progress.advance(rec.status, sku)
            if rec.status == "success":
                satisfied += 1
        return records

    all_records: list[ImageGenerationRecord] = []
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {
            executor.submit(process_sku, sku, srows): sku
            for sku, srows in by_sku.items()
        }
        for future in as_completed(futures):
            all_records.extend(future.result())
    return all_records


def process_one(
    index: int,
    total: int,
    row: dict[str, Any],
    prompt_map: dict[str, dict[int, str]],
    config: MxapiGenerateImagesConfig,
    client: MxapiImageClient,
    checkpoint_store: CheckpointStore,
) -> ImageGenerationRecord:
    sku = row["sku"]
    image_name = row["image_name"]
    image_number = row.get("image_number")
    print(f"[{index}/{total}] 开始生成 | SKU={sku} | 图片={image_name}", flush=True)
    if config.prompt_mode == "fixed":
        # 固定提示词模式（主图）：直接取输入 Excel 的「生成提示词」列，不走 BUZZ image_plan
        prompt = (row.get("prompt") or "").strip()
        prompt_error_code = "PromptEmpty"
        prompt_error_msg = "未找到提示词（fixed 模式请检查输入 Excel 的「生成提示词」列）"
    else:
        prompt = prompt_map.get(sku, {}).get(image_number or -1, "")
        prompt_error_code = "PromptNotFound"
        prompt_error_msg = "未找到对应 image_plan prompt"
    if not prompt:
        record = build_error_record(row, prompt_error_code, prompt_error_msg)
        checkpoint_store.upsert(record)
        print(f"[{index}/{total}] 生成失败 | SKU={sku} | 错误={prompt_error_msg}", flush=True)
        return record
    if not row.get("reference_image"):
        record = build_error_record(row, "ReferenceImageMissing", "参考图片链接为空")
        checkpoint_store.upsert(record)
        print(f"[{index}/{total}] 生成失败 | SKU={sku} | 错误=参考图片链接为空", flush=True)
        return record

    task_id = (row.get("task_id") or None) if config.poll_existing_task_id else None
    submit_latency_ms = None
    try:
        if task_id:
            # 断点续跑：先用原 task_id 轮询，任务已完成则直接复用（覆盖超时/抖动恢复）
            try:
                poll_result, poll_count, wait_seconds = poll_until_done(client, task_id, config)
            except TaskFailedError as exc:
                if is_permanent_failure(str(exc)):
                    # 永久失败（内容安全拦截）：不重新提交，标记 failed_permanent，后续行（后位类型）自然补位
                    record = build_permanent_record(row, str(exc), task_id=task_id)
                    checkpoint_store.upsert(record)
                    print(
                        f"[{index}/{total}] 永久失败(安全策略) | SKU={sku} | 类型={row.get('image_type') or '-'} | 错误={short_error(str(exc))}",
                        flush=True,
                    )
                    return record
                # 原任务确认彻底失败（status=failed）→ 丢弃旧 task_id，重新提交生成
                print(
                    f"[{index}/{total}] 旧任务失败，重新生成 | SKU={sku} | 旧task_id={task_id} | 原因={short_error(str(exc))}",
                    flush=True,
                )
                task_id, submit_latency_ms = submit_with_retry(client, prompt, row["reference_image"], config)
                checkpoint_store.upsert(build_submitted_record(row, task_id, submit_latency_ms))
                print(f"[{index}/{total}] 已提交 | SKU={sku} | task_id={task_id}", flush=True)
                poll_result, poll_count, wait_seconds = poll_until_done(client, task_id, config)
        else:
            task_id, submit_latency_ms = submit_with_retry(client, prompt, row["reference_image"], config)
            checkpoint_store.upsert(build_submitted_record(row, task_id, submit_latency_ms))
            print(f"[{index}/{total}] 已提交 | SKU={sku} | task_id={task_id}", flush=True)
            poll_result, poll_count, wait_seconds = poll_until_done(client, task_id, config)
        image_urls = collect_image_urls(poll_result)
        if not image_urls:
            raise RuntimeError("结果中无图片URL")
        download_path = Path(config.download_dir) / f"{image_name}.png"
        file_size, used_url = download_with_retry(client, image_urls, download_path, config)
        save_raw_response(config, image_name, poll_result)
        record = ImageGenerationRecord(
            row_number=row["row_number"],
            sku=sku,
            image_name=image_name,
            image_type=row.get("image_type"),
            image_number=image_number,
            status="success",
            task_id=task_id,
            reference_image=row.get("reference_image"),
            generated_image_url=used_url,
            downloaded_path=str(download_path),
            file_size=file_size,
            error_message=None,
            retryable=False,
            submit_latency_ms=submit_latency_ms,
            poll_count=poll_count,
            total_wait_seconds=wait_seconds,
            created_at=datetime.now().isoformat(timespec="seconds"),
        )
        checkpoint_store.upsert(record)
        print(f"[{index}/{total}] 生成成功 | SKU={sku} | 文件={download_path.name}", flush=True)
        return record
    except PollTimeoutError as exc:
        # 超时≠失败：保留 task_id，标记 pending，待下次续跑按 task_id 重新查询收敛。
        record = build_pending_record(row, str(exc), task_id=task_id, submit_latency_ms=submit_latency_ms)
        checkpoint_store.upsert(record)
        print(f"[{index}/{total}] 轮询超时(保留task_id待续跑) | SKU={sku} | task_id={task_id}", flush=True)
        return record
    except Exception as exc:
        error_msg = str(exc)
        if is_permanent_failure(error_msg):
            record = build_permanent_record(row, error_msg, task_id=task_id, submit_latency_ms=submit_latency_ms)
            print(
                f"[{index}/{total}] 永久失败(安全策略) | SKU={sku} | 类型={row.get('image_type') or '-'} | 错误={short_error(error_msg)}",
                flush=True,
            )
        else:
            record = build_error_record(row, exc.__class__.__name__, error_msg, task_id=task_id, submit_latency_ms=submit_latency_ms)
            print(f"[{index}/{total}] 生成失败 | SKU={sku} | 错误={short_error(error_msg)}", flush=True)
        checkpoint_store.upsert(record)
        return record


def submit_with_retry(client: MxapiImageClient, prompt: str, reference_image: str, config: MxapiGenerateImagesConfig) -> tuple[str, int | None]:
    payload = {
        "prompt": prompt,
        "aspect_ratio": config.aspect_ratio,
        "quality": config.quality,
        "resolution": config.resolution,
        "reference_images": [reference_image],
    }
    last_error: Exception | None = None
    for attempt in range(1, config.max_submit_retries + 1):
        if config.submit_delay_seconds > 0:
            time.sleep(config.submit_delay_seconds)
        try:
            response, latency_ms = client.submit(payload)
            if response.get("code") != 200:
                raise RuntimeError(str(response.get("message") or response)[:1000])
            task_id = (response.get("data") or {}).get("task_id")
            if not task_id:
                raise RuntimeError("No task_id in submit response")
            return str(task_id), latency_ms
        except Exception as exc:
            last_error = exc
            if attempt >= config.max_submit_retries:
                break
            time.sleep(config.retry_delay_seconds * attempt)
    raise RuntimeError(f"submit failed: {last_error}")


class TaskFailedError(RuntimeError):
    """MXAPI 任务确认失败（data.status == 'failed'），区别于超时/网络异常。"""


class PollTimeoutError(RuntimeError):
    """本地轮询超时：仅表示本侧未在 max_wait_seconds 内等到结果，平台侧任务可能已成功。"""


def poll_until_done(client: MxapiImageClient, task_id: str, config: MxapiGenerateImagesConfig) -> tuple[dict[str, Any], int, int]:
    started = time.time()
    poll_count = 0
    last_payload: dict[str, Any] = {}
    while time.time() - started <= config.max_wait_seconds:
        time.sleep(config.poll_interval_seconds)
        poll_count += 1
        payload, _ = client.query(task_id)
        last_payload = payload
        if payload.get("code") != 200:
            raise RuntimeError(str(payload.get("message") or payload)[:1000])
        data = payload.get("data") or {}
        status = data.get("status")
        if status == "completed":
            return payload, poll_count, int(time.time() - started)
        if status == "failed":
            # 真实错误在 error_msg 字段（部分服务返回 error），一并兼容读取
            error_msg = str(data.get("error_msg") or data.get("error") or "task failed")
            raise TaskFailedError(error_msg)
    raise PollTimeoutError(f"poll timeout after {config.max_wait_seconds}s: {last_payload}")


def collect_image_urls(payload: dict[str, Any]) -> list[str]:
    """从 MXAPI 返回中提取所有图片地址（source_images / proxy_images / images），去重保序。"""
    result = ((payload.get("data") or {}).get("result") or {})
    urls: list[str] = []
    for key in ("source_images", "proxy_images", "images"):
        values = result.get(key) or []
        if isinstance(values, list):
            urls.extend(str(item) for item in values if item)
    seen: set[str] = set()
    ordered: list[str] = []
    for url in urls:
        if url not in seen:
            seen.add(url)
            ordered.append(url)
    return ordered


def short_error(message: str, limit: int = 120) -> str:
    text = " ".join(str(message).split())
    if "<!DOCTYPE html>" in text or "<html" in text.lower():
        if "HTTP 520" in text:
            return "HTTP 520 Cloudflare HTML 错误页，上游/中转站异常"
        if "HTTP 504" in text:
            return "HTTP 504 网关超时"
        return "HTTP HTML 错误页，上游/中转站异常"
    return text if len(text) <= limit else text[:limit] + "..."


def download_with_retry(client: MxapiImageClient, urls: list[str], path: Path, config: MxapiGenerateImagesConfig) -> tuple[int, str]:
    """逐个地址尝试下载，某地址全部重试失败后自动切换到下一个地址。返回 (文件大小, 实际使用的 URL)。"""
    last_error: Exception | None = None
    tried: list[str] = []
    for url_index, url in enumerate(urls):
        per_url_error: Exception | None = None
        for attempt in range(1, config.max_download_retries + 1):
            try:
                size = client.download(url, path, timeout_seconds=config.download_timeout_seconds)
                return size, url
            except Exception as exc:
                per_url_error = exc
                if attempt >= config.max_download_retries:
                    break
                time.sleep(config.retry_delay_seconds * attempt)
        tried.append(f"url#{url_index}({str(url)[:60]}): {per_url_error}")
        last_error = per_url_error
    raise RuntimeError(f"download failed after trying {len(urls)} url(s): {last_error}; " + " | ".join(tried))


def save_raw_response(config: MxapiGenerateImagesConfig, image_name: str, payload: dict[str, Any]) -> None:
    if not config.raw_responses_dir:
        return
    path = Path(config.raw_responses_dir) / f"{image_name}.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")



def build_submitted_record(
    row: dict[str, Any],
    task_id: str,
    submit_latency_ms: int | None,
) -> ImageGenerationRecord:
    return ImageGenerationRecord(
        row_number=row["row_number"],
        sku=row["sku"],
        image_name=row["image_name"],
        image_type=row.get("image_type"),
        image_number=row.get("image_number"),
        status="submitted",
        task_id=task_id,
        reference_image=row.get("reference_image"),
        generated_image_url=None,
        downloaded_path=None,
        file_size=None,
        error_message=None,
        retryable=True,
        submit_latency_ms=submit_latency_ms,
        poll_count=0,
        total_wait_seconds=None,
        created_at=datetime.now().isoformat(timespec="seconds"),
    )
def build_error_record(
    row: dict[str, Any],
    error_code: str,
    error_message: str,
    task_id: str | None = None,
    submit_latency_ms: int | None = None,
) -> ImageGenerationRecord:
    return ImageGenerationRecord(
        row_number=row["row_number"],
        sku=row["sku"],
        image_name=row["image_name"],
        image_type=row.get("image_type"),
        image_number=row.get("image_number"),
        status="failed",
        task_id=task_id or row.get("task_id") or None,
        reference_image=row.get("reference_image"),
        generated_image_url=None,
        downloaded_path=None,
        file_size=None,
        error_message=f"{error_code}: {error_message}",
        retryable=error_code not in {"PromptNotFound", "ReferenceImageMissing"},
        submit_latency_ms=submit_latency_ms,
        poll_count=0,
        total_wait_seconds=None,
        created_at=datetime.now().isoformat(timespec="seconds"),
    )


PERMANENT_FAILURE_MARKERS = ("safety policy", "content safety", "content_safety", "sensitive")


def is_permanent_failure(error_message: str) -> bool:
    """判断是否永久失败（内容安全拦截等，重试无意义）。默认宽松：仅明确命中才判永久。"""
    msg = str(error_message).lower()
    return any(marker in msg for marker in PERMANENT_FAILURE_MARKERS)


def build_permanent_record(
    row: dict[str, Any],
    error_message: str,
    task_id: str | None = None,
    submit_latency_ms: int | None = None,
) -> ImageGenerationRecord:
    """永久失败记录：内容安全拦截等，重试/兜底均无意义；跨重启不再重新提交该类型。"""
    return ImageGenerationRecord(
        row_number=row["row_number"],
        sku=row["sku"],
        image_name=row["image_name"],
        image_type=row.get("image_type"),
        image_number=row.get("image_number"),
        status="failed_permanent",
        task_id=task_id or row.get("task_id") or None,
        reference_image=row.get("reference_image"),
        generated_image_url=None,
        downloaded_path=None,
        file_size=None,
        error_message=f"PermanentFailure: {error_message}",
        retryable=False,
        submit_latency_ms=submit_latency_ms,
        poll_count=0,
        total_wait_seconds=None,
        created_at=datetime.now().isoformat(timespec="seconds"),
    )


def build_pending_record(
    row: dict[str, Any],
    error_message: str,
    task_id: str | None = None,
    submit_latency_ms: int | None = None,
) -> ImageGenerationRecord:
    """轮询超时专用记录：状态 pending，保留 task_id，待下次续跑按 task_id 重新查询。不判失败。"""
    return ImageGenerationRecord(
        row_number=row["row_number"],
        sku=row["sku"],
        image_name=row["image_name"],
        image_type=row.get("image_type"),
        image_number=row.get("image_number"),
        status="pending",
        task_id=task_id or row.get("task_id") or None,
        reference_image=row.get("reference_image"),
        generated_image_url=None,
        downloaded_path=None,
        file_size=None,
        error_message=f"PollTimeout: {error_message}",
        retryable=True,
        submit_latency_ms=submit_latency_ms,
        poll_count=0,
        total_wait_seconds=None,
        created_at=datetime.now().isoformat(timespec="seconds"),
    )


def build_blocked_record(row: dict[str, Any]) -> ImageGenerationRecord:
    """依赖未满足（如主图未成功）：不提交 MXAPI、不消耗积分；非终态(retryable)，续跑会重新判定，主图成功后自动补生成。"""
    return ImageGenerationRecord(
        row_number=row["row_number"],
        sku=row["sku"],
        image_name=row["image_name"],
        image_type=row.get("image_type"),
        image_number=row.get("image_number"),
        status="blocked",
        task_id=None,
        reference_image=row.get("reference_image"),
        generated_image_url=None,
        downloaded_path=None,
        file_size=None,
        error_message="DependencyNotMet: 主图未成功生成（仅主图成功才生成副图）",
        retryable=True,
        submit_latency_ms=None,
        poll_count=0,
        total_wait_seconds=None,
        created_at=datetime.now().isoformat(timespec="seconds"),
    )


def build_skipped_record(row: dict[str, Any]) -> ImageGenerationRecord:
    """该 SKU 已达目标张数，跳过剩余类型行（不生成、不判失败，续跑也不再处理）。"""
    return ImageGenerationRecord(
        row_number=row["row_number"],
        sku=row["sku"],
        image_name=row["image_name"],
        image_type=row.get("image_type"),
        image_number=row.get("image_number"),
        status="skipped",
        task_id=None,
        reference_image=row.get("reference_image"),
        generated_image_url=None,
        downloaded_path=None,
        file_size=None,
        error_message="skipped: desired_count satisfied",
        retryable=False,
        submit_latency_ms=None,
        poll_count=0,
        total_wait_seconds=None,
        created_at=datetime.now().isoformat(timespec="seconds"),
    )


def merge_records(existing_rows: list[dict[str, Any]], new_records: list[ImageGenerationRecord], source_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged = {f"{row.get('sku')}::{row.get('image_name')}": row for row in existing_rows if row.get("sku") and row.get("image_name")}
    for record in new_records:
        merged[f"{record.sku}::{record.image_name}"] = asdict(record)
    ordered = []
    seen = set()
    for row in source_rows:
        key = row_key(row)
        if key in merged:
            current = dict(merged[key])
            current["row_number"] = row["row_number"]
            current["reference_image"] = row.get("reference_image")
            current["image_number"] = row.get("image_number")
            # 历史 checkpoint 记录可能缺 image_type（新增列之前生成）：以 Excel 行补齐，
            # 否则 06 的缺失类型清单会把“已成功但类型缺失”的图误判为未生成。
            current["image_type"] = row.get("image_type")
            ordered.append(current)
            seen.add(key)
    return ordered


def read_jsonl_if_exists(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                # 追加写崩溃可能残留半行：跳过，同 key 的后续行仍会覆盖旧值
                continue
    return rows


def write_jsonl_rows(rows: list[dict[str, Any]], path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_excel(workbook, sheet, headers: dict[str, int], rows: list[dict[str, Any]], config: MxapiGenerateImagesConfig) -> None:
    row_numbers_by_key = current_sheet_row_numbers(sheet, headers, config)
    for row in rows:
        row_number = row_numbers_by_key.get(record_key(row))
        if not row_number:
            continue
        status = row.get("status")
        status_text = (
            "成功" if status == "success"
            else "超时待定" if status == "pending"
            else "跳过" if status == "skipped"
            else "永久失败" if status == "failed_permanent"
            else "依赖未满足" if status == "blocked"
            else "失败"
        )
        sheet.cell(row_number, headers[config.columns["status"]], value=status_text)
        sheet.cell(row_number, headers[config.columns["task_id"]], value=row.get("task_id"))
    output_path = Path(config.output_excel_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"图片结果日志: {config.output_results_path}")
    print(f"图片结果Excel: {config.output_excel_path}")


def current_sheet_row_numbers(sheet, headers: dict[str, int], config: MxapiGenerateImagesConfig) -> dict[str, int]:
    sku_col = headers[config.columns["sku"]]
    image_name_col = headers[config.columns["image_name"]]
    row_numbers: dict[str, int] = {}
    for row_number in range(2, sheet.max_row + 1):
        sku = cell_text(sheet, row_number, sku_col)
        image_name = cell_text(sheet, row_number, image_name_col)
        if not sku or not image_name:
            continue
        row_numbers.setdefault(f"{sku}::{image_name}", row_number)
    return row_numbers


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate images with MXAPI.")
    parser.add_argument("--config", required=True)
    parser.add_argument("--max-records", type=int, default=None)
    parser.add_argument("--concurrency", type=int, default=None)
    args = parser.parse_args()
    config = load_config(args.config)
    if args.max_records is not None:
        config.max_records = args.max_records
    if args.concurrency is not None:
        config.concurrency = args.concurrency
    records = run(config)
    success_count = sum(1 for item in records if item.status == "success")
    print(f"图片生成汇总: {len(records)} | 成功: {success_count} | 失败: {len(records) - success_count}")


if __name__ == "__main__":
    main()



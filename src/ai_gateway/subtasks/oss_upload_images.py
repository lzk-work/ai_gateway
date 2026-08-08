"""Upload generated images to Aliyun OSS with per-image checkpointing."""

from __future__ import annotations

import argparse
import json
import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ai_gateway.clients.aliyun_oss_client import (
    AliyunOssClient,
    build_public_url_without_client,
    load_aliyun_oss_config,
)
from ai_gateway.config.loader import load_local_env


@dataclass(slots=True)
class OssUploadConfig:
    name: str
    project_root: str
    input_excel_path: str
    input_sheet_name: str
    download_dir: str
    output_excel_path: str
    output_results_path: str
    checkpoint_path: str
    columns: dict[str, str]
    oss_prefix: str
    key_template: str
    url_template: str | None = None
    overwrite: bool = True
    max_records: int | None = None
    concurrency: int = 5
    batch_size: int = 500
    skip_success: bool = True


@dataclass(slots=True)
class OssUploadRecord:
    row_number: int
    sku: str
    image_name: str
    status: str
    local_path: str
    oss_key: str
    oss_url: str | None
    file_size: int | None
    error_message: str | None
    retryable: bool
    created_at: str


class CheckpointStore:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        self.lock = threading.Lock()
        self.records: dict[str, dict[str, Any]] = {}
        for row in read_jsonl_if_exists(self.path):
            key = record_key(row)
            if key:
                self.records[key] = row

    def rows(self) -> list[dict[str, Any]]:
        with self.lock:
            return list(self.records.values())

    def upsert(self, record: OssUploadRecord) -> None:
        with self.lock:
            self.records[record_key(asdict(record))] = asdict(record)
            self.flush_locked()

    def flush_locked(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = self.path.with_suffix(self.path.suffix + ".tmp")
        with temp_path.open("w", encoding="utf-8") as handle:
            for row in self.records.values():
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
        temp_path.replace(self.path)


def find_project_root(path: Path) -> Path:
    for candidate in [path.parent, *path.parents]:
        if (candidate / "src" / "ai_gateway").exists() and (candidate / "configs").exists():
            return candidate
    raise RuntimeError(f"Cannot find project root from config path: {path}")


def load_config(path: str | Path) -> OssUploadConfig:
    path = Path(path)
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    project_root = find_project_root(path.resolve())
    oss = data.get("oss", {})
    limits = data.get("limits", {})
    resume = data.get("resume", {})
    return OssUploadConfig(
        name=data["name"],
        project_root=str(project_root),
        input_excel_path=data["input"]["image_result_excel_path"],
        input_sheet_name=data["input"].get("sheet_name", "Sheet1"),
        download_dir=data["input"]["download_dir"],
        output_excel_path=data["output"]["excel_path"],
        output_results_path=data["output"]["results_path"],
        checkpoint_path=data["output"]["checkpoint_path"],
        columns=data["columns"],
        oss_prefix=oss.get("prefix", "walmart").strip("/"),
        key_template=oss.get("key_template", "walmart/{sku}/{image_name}.png"),
        url_template=oss.get("url_template"),
        overwrite=bool(oss.get("overwrite", True)),
        concurrency=int(limits.get("max_workers", 5)),
        batch_size=int(limits.get("batch_size", 500)),
        skip_success=bool(resume.get("skip_success", True)),
    )


def run(config: OssUploadConfig) -> list[OssUploadRecord]:
    rows, workbook, sheet, headers = load_work_rows(config)
    checkpoint = CheckpointStore(config.checkpoint_path)
    existing = checkpoint.rows()
    completed = completed_keys(existing) if config.skip_success else set()
    pending = [row for row in rows if row_key(row) not in completed]
    if config.max_records and config.max_records > 0:
        pending = pending[: config.max_records]

    print("\n=== 05 上传 OSS ===", flush=True)
    print(
        f"图片总数: {len(rows)} | 已成功跳过: {len(rows) - len(pending)} | 本次待上传: {len(pending)}",
        flush=True,
    )
    client = AliyunOssClient(load_aliyun_oss_config(config.project_root))
    records = process_rows(pending, config, client, checkpoint)
    merged = merge_records(checkpoint.rows(), records, rows)
    write_jsonl_rows(merged, config.output_results_path)
    write_excel(workbook, sheet, headers, merged, config)
    return records


def preview(config: OssUploadConfig) -> None:
    if not Path(config.input_excel_path).exists():
        print("\n=== 05 上传 OSS | 试运行 ===")
        print("试运行: 不连接 OSS，不上传，不写 JSONL/Excel")
        print(f"OSS 上传输入 Excel 不存在: {config.input_excel_path}")
        print("请先完成 03 图片生成下载，或确认当前入参文件名对应的批次目录是否正确。")
        return
    rows, _, _, _ = load_work_rows(config)
    existing = read_jsonl_if_exists(config.checkpoint_path)
    completed = completed_keys(existing) if config.skip_success else set()
    pending = [row for row in rows if row_key(row) not in completed]
    if config.max_records and config.max_records > 0:
        selected = pending[: config.max_records]
    else:
        selected = pending

    env_values = load_local_env(Path(config.project_root) / "configs" / "local.env")
    bucket = os.environ.get("ALIYUN_OSS_BUCKET") or env_values.get("ALIYUN_OSS_BUCKET") or "<bucket>"
    endpoint = os.environ.get("ALIYUN_OSS_ENDPOINT") or env_values.get("ALIYUN_OSS_ENDPOINT") or "<endpoint>"
    default_prefix = os.environ.get("ALIYUN_OSS_DEFAULT_PREFIX") or env_values.get("ALIYUN_OSS_DEFAULT_PREFIX") or "images"

    print("\n=== 05 上传 OSS | 试运行 ===")
    print("试运行: 不连接 OSS，不上传，不写 JSONL/Excel")
    print(f"输入Excel: {config.input_excel_path}")
    print(f"图片目录: {config.download_dir}")
    print(f"checkpoint: {config.checkpoint_path}")
    print(f"图片总数: {len(rows)} | 已成功跳过: {len(rows) - len(pending)} | 本次将上传: {len(selected)}")
    print(f"并发: {config.concurrency} | batch_size: {config.batch_size} | overwrite: {config.overwrite}")
    if selected:
        print("样例:")
        for index, row in enumerate(selected[:5], start=1):
            oss_key = row["oss_key"]
            url = build_public_url_without_client(bucket, endpoint, default_prefix, oss_key)
            print(f"  [{index}] SKU={row['sku']} | 图片={Path(row['local_path']).name} | OSS={url}")


def load_work_rows(config: OssUploadConfig):
    input_path = Path(config.input_excel_path)
    if not input_path.exists():
        raise RuntimeError(f"OSS 上传输入 Excel 不存在: {input_path}")
    workbook = load_workbook(input_path)
    sheet = workbook[config.input_sheet_name]
    headers = header_map(sheet)
    required = [
        config.columns["sku"],
        config.columns["image_name"],
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError("Missing OSS upload input headers: " + ", ".join(missing))

    rows: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    for row_number in range(2, sheet.max_row + 1):
        sku = cell_text(sheet, row_number, headers[config.columns["sku"]])
        image_name = cell_text(sheet, row_number, headers[config.columns["image_name"]])
        if not sku or not image_name:
            continue
        download_status_col = config.columns.get("download_status")
        if download_status_col and download_status_col in headers:
            download_status = cell_text(sheet, row_number, headers[download_status_col])
            if download_status != "成功":
                continue
        unique_key = f"{sku}::{image_name}"
        if unique_key in seen_keys:
            continue
        seen_keys.add(unique_key)
        local_path = Path(config.download_dir) / image_file_name(image_name)
        oss_key = render_template(config.key_template, sku=sku, image_name=Path(image_file_name(image_name)).stem)
        rows.append(
            {
                "row_number": row_number,
                "sku": sku,
                "image_name": image_name,
                "local_path": str(local_path),
                "oss_key": oss_key,
            }
        )
    return rows, workbook, sheet, headers


def process_rows(
    rows: list[dict[str, Any]],
    config: OssUploadConfig,
    client: AliyunOssClient,
    checkpoint: CheckpointStore,
) -> list[OssUploadRecord]:
    if not rows:
        return []
    concurrency = max(int(config.concurrency or 1), 1)
    concurrency = min(concurrency, len(rows))
    print(f"OSS上传并发: {concurrency}", flush=True)
    if concurrency == 1:
        return [process_one(index, len(rows), row, config, client, checkpoint) for index, row in enumerate(rows, start=1)]
    results: dict[int, OssUploadRecord] = {}
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {
            executor.submit(process_one, index, len(rows), row, config, client, checkpoint): index
            for index, row in enumerate(rows, start=1)
        }
        for future in as_completed(futures):
            index = futures[future]
            results[index] = future.result()
    return [results[index] for index in sorted(results)]


def process_one(
    index: int,
    total: int,
    row: dict[str, Any],
    config: OssUploadConfig,
    client: AliyunOssClient,
    checkpoint: CheckpointStore,
) -> OssUploadRecord:
    sku = row["sku"]
    local_path = Path(row["local_path"])
    print(f"[{index}/{total}] 开始上传 | SKU={sku} | 图片={local_path.name}", flush=True)
    if not local_path.is_file():
        record = build_record(row, "missing_file", None, f"本地图片不存在: {local_path}", retryable=False)
        checkpoint.upsert(record)
        print(f"[{index}/{total}] 上传失败 | SKU={sku} | 错误=本地图片不存在", flush=True)
        return record

    result = client.upload_file(local_path, row["oss_key"], overwrite=config.overwrite)
    if result.get("success"):
        status = "skipped" if result.get("skipped") else "success"
        url = client.public_url(row["oss_key"])
        record = build_record(row, status, url, None, retryable=False, file_size=result.get("size"))
        checkpoint.upsert(record)
        label = "已存在跳过" if status == "skipped" else "上传成功"
        print(f"[{index}/{total}] {label} | SKU={sku} | URL={url}", flush=True)
        return record

    record = build_record(row, "failed", None, str(result.get("error") or "upload failed"), retryable=True)
    checkpoint.upsert(record)
    print(f"[{index}/{total}] 上传失败 | SKU={sku} | 错误={short_error(record.error_message or '')}", flush=True)
    return record


def build_record(
    row: dict[str, Any],
    status: str,
    oss_url: str | None,
    error_message: str | None,
    retryable: bool,
    file_size: int | None = None,
) -> OssUploadRecord:
    return OssUploadRecord(
        row_number=row["row_number"],
        sku=row["sku"],
        image_name=row["image_name"],
        status=status,
        local_path=row["local_path"],
        oss_key=row["oss_key"],
        oss_url=oss_url,
        file_size=file_size,
        error_message=error_message,
        retryable=retryable,
        created_at=datetime.now().isoformat(timespec="seconds"),
    )


def merge_records(existing_rows: list[dict[str, Any]], new_records: list[OssUploadRecord], source_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged = {record_key(row): row for row in existing_rows if record_key(row)}
    for record in new_records:
        merged[record_key(asdict(record))] = asdict(record)
    ordered = []
    seen = set()
    for row in source_rows:
        key = row_key(row)
        if key in merged:
            ordered.append(merged[key])
            seen.add(key)
    return ordered


def write_excel(workbook, sheet, headers: dict[str, int], rows: list[dict[str, Any]], config: OssUploadConfig) -> None:
    result_columns = ["OSS上传状态", "OSS上传URL", "OSS上传错误", "OSS上传时间"]
    for column_name in result_columns:
        if column_name not in headers:
            headers[column_name] = sheet.max_column + 1
            sheet.cell(row=1, column=headers[column_name], value=column_name)

    for row in rows:
        row_number = row.get("row_number")
        if not isinstance(row_number, int):
            continue
        sheet.cell(row_number, headers["OSS上传状态"], value=display_status(str(row.get("status") or "")))
        sheet.cell(row_number, headers["OSS上传URL"], value=row.get("oss_url"))
        sheet.cell(row_number, headers["OSS上传错误"], value=row.get("error_message"))
        sheet.cell(row_number, headers["OSS上传时间"], value=row.get("created_at"))

    output_path = Path(config.output_excel_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"OSS结果日志: {config.output_results_path}")
    print(f"OSS结果Excel: {config.output_excel_path}")


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def cell_text(sheet, row: int, column: int) -> str:
    value = sheet.cell(row, column).value
    return "" if value is None else str(value).strip()


def image_file_name(image_name: str) -> str:
    suffix = Path(image_name).suffix
    return image_name if suffix else f"{image_name}.png"


def render_template(template: str, **values: str) -> str:
    rendered = template
    for key, value in values.items():
        rendered = rendered.replace("{" + key + "}", str(value))
    return rendered.replace("\\", "/").lstrip("/")


def row_key(row: dict[str, Any]) -> str:
    return f"{row['sku']}::{row['image_name']}"


def record_key(row: dict[str, Any]) -> str:
    sku = row.get("sku")
    image_name = row.get("image_name")
    return f"{sku}::{image_name}" if sku and image_name else ""


def completed_keys(rows: list[dict[str, Any]]) -> set[str]:
    return {
        record_key(row)
        for row in rows
        if row.get("status") in {"success", "skipped"} and record_key(row)
    }


def read_jsonl_if_exists(path: str | Path) -> list[dict[str, Any]]:
    path = Path(path)
    if not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def write_jsonl_rows(rows: list[dict[str, Any]], path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def display_status(status: str) -> str:
    return {
        "success": "成功",
        "skipped": "跳过",
        "failed": "失败",
        "missing_file": "本地图片不存在",
    }.get(status, status)


def short_error(message: str, limit: int = 120) -> str:
    text = " ".join(str(message).split())
    return text if len(text) <= limit else text[:limit] + "..."


def main() -> None:
    parser = argparse.ArgumentParser(description="Upload generated images to Aliyun OSS.")
    parser.add_argument("--config", required=True)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--max-records", type=int, default=None)
    parser.add_argument("--concurrency", type=int, default=None)
    args = parser.parse_args()
    config = load_config(args.config)
    if args.max_records is not None:
        config.max_records = args.max_records
    if args.concurrency is not None:
        config.concurrency = args.concurrency
    if args.dry_run:
        preview(config)
        return
    records = run(config)
    success_count = sum(1 for item in records if item.status in {"success", "skipped"})
    print(f"OSS上传汇总: {len(records)} | 成功/跳过: {success_count} | 失败: {len(records) - success_count}")


if __name__ == "__main__":
    main()

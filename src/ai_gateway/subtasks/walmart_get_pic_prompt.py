"""Build Walmart image prompt tasks from an Excel file and prompt template."""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from ai_gateway.models import ImageInput
from ai_gateway.subtasks.template_renderer import list_placeholders, render_template


@dataclass(slots=True)
class WalmartPromptConfig:
    name: str
    input_excel: str
    prompt_template_path: str
    output_path: str
    sheet_name: str | None = None
    batch_id: str | None = None
    task_id_column: str = "开发SKU"
    title_column: str = "标题"
    bullet_column: str = "五点"
    image_columns: list[str] = field(default_factory=lambda: ["主图"])
    context_policy: dict[str, Any] = field(
        default_factory=lambda: {
            "mode": "stateless",
            "include_history": False,
            "history_turns": 0,
        }
    )
    limits: dict[str, Any] = field(
        default_factory=lambda: {
            "max_prompt_chars": 60000,
            "warn_prompt_chars": 40000,
            "max_images_per_task": 1,
            "allowed_image_types": ["jpg", "jpeg", "png", "webp"],
            "image_detail": "auto",
        }
    )
    placeholder_mapping: dict[str, str] = field(
        default_factory=lambda: {"产品标题": "标题", "产品五点": "五点"}
    )
    include_row_metadata: bool = True


@dataclass(slots=True)
class WalmartPromptRecord:
    task_id: str
    batch_id: str
    sku: str
    row_number: int
    prompt_text: str
    image_urls: list[str]
    missing_placeholders: list[str]
    precheck_status: str
    precheck_warnings: list[str]
    precheck_errors: list[str]
    source_payload: dict[str, Any]
    next_task_payload: dict[str, Any]
    created_at: str


def load_config(path: str | Path) -> WalmartPromptConfig:
    data = json.loads(Path(path).read_text(encoding="utf-8-sig"))
    if "input" in data or "output" in data or "columns" in data:
        data = {
            "name": data["name"],
            "input_excel": data["input"]["excel_path"],
            "sheet_name": data["input"].get("sheet_name"),
            "prompt_template_path": data["input"]["prompt_template_path"],
            "output_path": data["output"]["prompt_tasks_path"],
            "batch_id": data.get("batch_id"),
            "task_id_column": data.get("columns", {}).get("task_id", "开发SKU"),
            "title_column": data.get("columns", {}).get("title", "标题"),
            "bullet_column": data.get("columns", {}).get("bullets", "五点"),
            "image_columns": data.get("columns", {}).get("images", ["主图"]),
            "context_policy": data.get("context_policy", {}),
            "limits": data.get("limits", {}),
            "placeholder_mapping": data.get("placeholder_mapping", {}),
            "include_row_metadata": data.get("include_row_metadata", True),
        }
    return WalmartPromptConfig(**data)


def run(config: WalmartPromptConfig) -> list[WalmartPromptRecord]:
    rows = read_excel_rows(config.input_excel, config.sheet_name)
    template = Path(config.prompt_template_path).read_text(encoding="utf-8-sig")
    placeholders = list_placeholders(template)
    records: list[WalmartPromptRecord] = []
    batch_id = config.batch_id or datetime.now().strftime("walmart_pic_prompt_%Y%m%d_%H%M%S")

    for row_number, row in rows:
        if _is_empty_row(row):
            continue
        rendered_prompt, missing = render_template(
            template,
            row,
            placeholder_mapping=config.placeholder_mapping,
        )
        image_urls = [
            str(row.get(column)).strip()
            for column in config.image_columns
            if row.get(column)
        ]
        precheck_status, warnings, errors = precheck_task(
            rendered_prompt,
            image_urls,
            config.limits,
        )
        sku = str(row.get(config.task_id_column) or f"row-{row_number}").strip()
        task_id = f"{batch_id}:{sku}"
        source_payload = {
            "sku": sku,
            "title": row.get(config.title_column),
            "bullets": row.get(config.bullet_column),
            "placeholder_names": placeholders,
        }
        if config.include_row_metadata:
            source_payload["row"] = row

        records.append(
            WalmartPromptRecord(
                task_id=task_id,
                batch_id=batch_id,
                sku=sku,
                row_number=row_number,
                prompt_text=rendered_prompt,
                image_urls=image_urls,
                missing_placeholders=missing,
                precheck_status=precheck_status,
                precheck_warnings=warnings,
                precheck_errors=errors,
                source_payload=source_payload,
                next_task_payload={
                    "task_id": task_id,
                    "batch_id": batch_id,
                    "prompt": rendered_prompt,
                    "images": [
                        asdict(ImageInput(type="url", value=url)) for url in image_urls
                    ],
                    "metadata": {
                        "subtask": config.name,
                        "sku": sku,
                        "source_row_number": row_number,
                    },
                    "context_policy": config.context_policy,
                    "limits": config.limits,
                    "precheck_status": precheck_status,
                    "precheck_warnings": warnings,
                    "precheck_errors": errors,
                },
                created_at=datetime.now().isoformat(timespec="seconds"),
            )
        )

    write_records(records, config.output_path)
    return records


def read_excel_rows(
    path: str | Path,
    sheet_name: str | None = None,
) -> list[tuple[int, dict[str, Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to read Excel files.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    try:
        header_values = next(row_iter)
    except StopIteration:
        return []

    headers = [str(value).strip() if value is not None else "" for value in header_values]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(row_iter, start=2):
        row = {
            headers[index]: value
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        rows.append((row_number, row))
    return rows


def precheck_task(
    prompt_text: str,
    image_urls: list[str],
    limits: dict[str, Any],
) -> tuple[str, list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    prompt_len = len(prompt_text)
    warn_prompt_chars = int(limits.get("warn_prompt_chars", 0) or 0)
    max_prompt_chars = int(limits.get("max_prompt_chars", 0) or 0)
    max_images = int(limits.get("max_images_per_task", 0) or 0)
    allowed_types = {
        str(item).lower().lstrip(".")
        for item in limits.get("allowed_image_types", [])
    }

    if warn_prompt_chars and prompt_len > warn_prompt_chars:
        warnings.append(f"prompt length {prompt_len} exceeds warning threshold {warn_prompt_chars}")
    if max_prompt_chars and prompt_len > max_prompt_chars:
        errors.append(f"prompt length {prompt_len} exceeds max {max_prompt_chars}")
    if max_images and len(image_urls) > max_images:
        errors.append(f"image count {len(image_urls)} exceeds max {max_images}")
    if not image_urls:
        errors.append("missing image url")

    for url in image_urls:
        suffix = url.split("?", 1)[0].rsplit(".", 1)[-1].lower() if "." in url else ""
        if allowed_types and suffix and suffix not in allowed_types:
            warnings.append(f"image url extension '{suffix}' is not in allowed types")

    return ("failed" if errors else "passed"), warnings, errors


def write_records(records: list[WalmartPromptRecord], output_path: str | Path) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".csv":
        _write_csv(records, path)
    else:
        _write_jsonl(records, path)


def _write_jsonl(records: list[WalmartPromptRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(asdict(record), ensure_ascii=False) + "\n")


def _write_csv(records: list[WalmartPromptRecord], path: Path) -> None:
    rows = []
    for record in records:
        row = asdict(record)
        for key in ("image_urls", "missing_placeholders", "source_payload", "next_task_payload"):
            row[key] = json.dumps(row[key], ensure_ascii=False)
        rows.append(row)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _is_empty_row(row: dict[str, Any]) -> bool:
    return not any(value not in (None, "") for value in row.values())


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Walmart image prompt tasks.")
    parser.add_argument("--config", required=True, help="Path to subtask config JSON.")
    args = parser.parse_args()
    records = run(load_config(args.config))
    print(f"generated={len(records)}")
    if records:
        print(f"batch_id={records[0].batch_id}")


if __name__ == "__main__":
    main()


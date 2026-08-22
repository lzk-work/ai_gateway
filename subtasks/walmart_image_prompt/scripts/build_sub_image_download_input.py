"""Build sub-image download input workbook for successful Walmart SKUs.

支持两种模式：
1. 默认（不配置 image_selection）：每个成功 SKU 生成 image_count 张副图行（new_sub1..N）。
2. 按图片类型选图（总配置 config.json 的 image_selection）：
   - image_type_order：6 种 image_type 按优先级排列，全部产出行作为候选；
   - desired_count：每个 SKU 的目标张数（03 按序处理、够数即停；某类型失败时后位类型自然补位）；
   行内写入「图片类型」列，供 03 阶段按序生成与续跑。
"""

from __future__ import annotations

import json
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ai_gateway.validators.result_validator import extract_json

TASK_ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = TASK_ROOT / "stages" / "build_sub_image_download_input" / "config.json"
TOTAL_CONFIG_PATH = TASK_ROOT / "config.json"


def header_map(worksheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def read_jsonl(path: str | Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    path = Path(path)
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def successful_skus(results_path: str | Path) -> set[str]:
    skus: set[str] = set()
    for row in read_jsonl(results_path):
        full_output_path = row.get("full_output_path")
        if full_output_path:
            text_path = Path(full_output_path)
            if not text_path.exists():
                continue
            _, _, validation_error = _inspect(text_path.read_text(encoding="utf-8", errors="replace"))
            if validation_error:
                continue
        elif row.get("status") != "success" or row.get("validation_status") != "passed":
            continue
        sku = row.get("sku")
        if sku:
            skus.add(str(sku).strip())
    return skus


def _inspect(text: str):
    # 延迟导入，避免无谓依赖；与 walmart_call_prompt_model.inspect_result_text 行为一致
    try:
        from ai_gateway.subtasks.walmart_call_prompt_model import inspect_result_text
        return inspect_result_text(text)
    except Exception:
        return None, None, None


def load_sku_type_maps(results_path: str | Path) -> dict[str, dict[str, int]]:
    """读取每个成功 SKU 的 BUZZ 方案，建立 {image_type(归一化) -> image_number} 映射。"""
    maps: dict[str, dict[str, int]] = {}
    for row in read_jsonl(results_path):
        if row.get("status") != "success" or row.get("validation_status") != "passed":
            continue
        sku = row.get("sku")
        full_output_path = row.get("full_output_path")
        if not sku or not full_output_path or not Path(full_output_path).exists():
            continue
        try:
            parsed, parse_error = extract_json(Path(full_output_path).read_text(encoding="utf-8"))
        except Exception:
            continue
        if parse_error or not isinstance(parsed, dict):
            continue
        type_map: dict[str, int] = {}
        for item in parsed.get("image_plan", []) or []:
            if not isinstance(item, dict):
                continue
            image_type = str(item.get("image_type") or "").strip()
            image_number = item.get("image_number")
            if image_type and image_number:
                type_map[image_type] = int(image_number)
        if type_map:
            maps[str(sku).strip()] = type_map
    return maps


def load_image_selection() -> dict[str, Any]:
    """从总配置读取 image_selection（单一数据源）。缺失或非法则返回空。"""
    if not TOTAL_CONFIG_PATH.exists():
        return {}
    try:
        data = json.loads(TOTAL_CONFIG_PATH.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}
    selection = data.get("image_selection") or {}
    if not isinstance(selection, dict):
        return {}
    return selection


def copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    for col_idx in range(1, worksheet.max_column + 1):
        source_cell = worksheet.cell(source_row, col_idx)
        target_cell = worksheet.cell(target_row, col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)


def require_headers(headers: dict[str, int], names: list[str], label: str) -> None:
    missing = [name for name in names if name not in headers]
    if missing:
        raise RuntimeError(f"Missing {label} headers: {', '.join(missing)}")


def ensure_image_type_column(worksheet, headers: dict[str, int], columns: dict[str, str]) -> tuple[dict[str, int], int | None]:
    """确保模板含「图片类型」列，返回 (最新 headers, 列号)。"""
    type_header = columns.get("template_image_type", "图片类型")
    if type_header in headers:
        return headers, headers[type_header]
    name_col = headers[columns["template_image_name"]]
    worksheet.insert_cols(name_col + 1)
    worksheet.cell(1, name_col + 1, type_header)
    new_headers = header_map(worksheet)
    return new_headers, new_headers.get(type_header)


def run(config: dict[str, Any]) -> None:
    input_config = config["input"]
    output_config = config["output"]
    columns = config["columns"]
    image_count = int(config.get("image_count", 6))

    # 选图配置（总配置 image_selection）
    selection = load_image_selection()
    image_type_order = [str(t).strip() for t in (selection.get("image_type_order") or []) if str(t).strip()]
    desired_count = selection.get("desired_count")
    desired_count = int(desired_count) if isinstance(desired_count, int) and desired_count > 0 else None
    use_selection = bool(image_type_order) and desired_count is not None
    type_maps = load_sku_type_maps(input_config["model_results_path"]) if use_selection else {}

    success_skus = successful_skus(input_config["model_results_path"])
    if not success_skus:
        raise RuntimeError("No successful SKUs found in model results.")

    source_wb = load_workbook(input_config["source_excel_path"], read_only=True, data_only=True)
    source_ws = source_wb[input_config["source_sheet_name"]]
    source_headers = header_map(source_ws)
    require_headers(
        source_headers,
        [columns["source_sku"], columns["source_main_image"]],
        "source Excel",
    )

    template_wb = load_workbook(input_config["template_path"])
    template_ws = template_wb[input_config["template_sheet_name"]]
    template_headers = header_map(template_ws)
    require_headers(
        template_headers,
        [
            columns["template_sku"],
            columns["template_reference_image"],
            columns["template_image_name"],
            columns["template_download_result"],
            columns["template_task_id"],
        ],
        "template Excel",
    )
    remove_columns_by_header(template_ws, ["OSS上传结果", "结果确认"])
    template_headers = header_map(template_ws)
    # 确保「图片类型」列存在（选图模式需要；非选图模式也无害保留）
    template_headers, type_col = ensure_image_type_column(template_ws, template_headers, columns)

    if template_ws.max_row > 1:
        template_ws.delete_rows(2, template_ws.max_row - 1)

    sku_col = source_headers[columns["source_sku"]]
    main_image_col = source_headers[columns["source_main_image"]]
    style_row = 2
    output_row = 2
    generated_skus: list[str] = []
    empty_sku_streak = 0
    stop_after_empty_sku_rows = int(config.get("stop_after_empty_sku_rows", 200))

    def write_row(sku: str, main_image: str, image_name: str, image_type: str) -> None:
        nonlocal output_row
        copy_row_style(template_ws, style_row, output_row)
        template_ws.cell(output_row, template_headers[columns["template_sku"]]).value = sku
        template_ws.cell(output_row, template_headers[columns["template_reference_image"]]).value = main_image
        template_ws.cell(output_row, template_headers[columns["template_image_name"]]).value = image_name
        template_ws.cell(output_row, template_headers[columns["template_download_result"]]).value = None
        template_ws.cell(output_row, template_headers[columns["template_task_id"]]).value = None
        if type_col is not None:
            template_ws.cell(output_row, type_col).value = image_type
        output_row += 1

    for row_idx in range(2, source_ws.max_row + 1):
        raw_sku = source_ws.cell(row_idx, sku_col).value
        if raw_sku is None:
            empty_sku_streak += 1
            if empty_sku_streak >= stop_after_empty_sku_rows:
                break
            continue
        sku = str(raw_sku).strip()
        if not sku:
            empty_sku_streak += 1
            if empty_sku_streak >= stop_after_empty_sku_rows:
                break
            continue
        empty_sku_streak = 0
        if sku not in success_skus:
            continue
        main_image = source_ws.cell(row_idx, main_image_col).value
        main_image = "" if main_image is None else str(main_image).strip()
        generated_skus.append(sku)

        if use_selection:
            chosen: list[tuple[int, str]] = []
            type_map = type_maps.get(sku, {})
            if type_map:
                # 全部类型都产出行（不止前 desired_count 种）。03 按 image_type_order 顺序处理、
                # 够 desired_count 即停；某类型失败时，列表后位类型（后续行）自然补位凑齐目标张数。
                for t in image_type_order:
                    number = type_map.get(t)
                    if number:
                        chosen.append((number, t))
            else:
                # 该 SKU 无方案类型映射：按序号退化生成全部候选行
                for image_no in range(1, len(image_type_order) + 1):
                    chosen.append((image_no, ""))
            for image_no, image_type in chosen:
                image_name = f"new_sub{image_no}_{sku}"
                write_row(sku, main_image, image_name, image_type)
        else:
            for image_no in range(1, image_count + 1):
                image_name = f"new_sub{image_no}_{sku}"
                write_row(sku, main_image, image_name, "")

    output_path = Path(output_config["excel_path"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    template_wb.save(output_path)
    mode = f"按类型选图(目标{desired_count}张, 候选{len(image_type_order)}种)" if use_selection else f"全量({image_count}张)"
    print(f"选图模式={mode}")
    print(f"successful_skus={len(generated_skus)}")
    print(f"generated_rows={output_row - 2}")
    print(f"output_path={output_path}")


def remove_columns_by_header(worksheet, header_names: list[str]) -> None:
    headers = header_map(worksheet)
    columns = sorted(
        [headers[name] for name in header_names if name in headers],
        reverse=True,
    )
    for column in columns:
        worksheet.delete_cols(column)


def main() -> None:
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8-sig"))
    run(config)


if __name__ == "__main__":
    main()

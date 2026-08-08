"""Batch-level progress statistics for the Walmart image prompt workflow."""

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from workflow_common import batch_name_from_input, batch_paths

from ai_gateway.subtasks.walmart_call_prompt_model import inspect_result_text
from final_image_result import preview_final_image_result_from_logs


IMAGE_COUNT = 6


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def excel_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [
        str(value).strip() if value is not None and str(value).strip() else ""
        for value in header_values
    ]
    rows: list[dict[str, str]] = []
    for values in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        row = {
            name: cell_text(values[index])
            for index, name in enumerate(headers)
            if name
        }
        if any(row.values()):
            rows.append(row)
    return rows


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def row_key(row: dict[str, Any]) -> str:
    return f"{row.get('sku')}::{row.get('image_name')}"


def image_row_key(row: dict[str, str]) -> str:
    return f"{row.get('SKU')}::{row.get('图片命名')}"


def model_success_is_valid(row: dict[str, Any]) -> tuple[bool, str | None]:
    if row.get("status") != "success" or row.get("validation_status") != "passed":
        return False, row.get("error_message") or row.get("status")
    full_output_path = row.get("full_output_path")
    if full_output_path:
        path = Path(full_output_path)
        if not path.exists():
            return False, "完整输出文件不存在"
        _, _, validation_error = inspect_result_text(path.read_text(encoding="utf-8", errors="replace"))
        if validation_error:
            return False, validation_error
    return True, None


def collect_stats(batch_name: str | None = None) -> dict[str, Any]:
    effective_batch = batch_name or batch_name_from_input()
    paths = batch_paths(effective_batch)

    prompt_rows = read_jsonl(paths["prompt_tasks"])
    model_rows = read_jsonl(paths["model_results"])
    image_input_rows = excel_rows(paths["image_input_excel"])
    image_rows = read_jsonl(paths["image_results"])
    upload_rows = read_jsonl(paths["oss_results"])

    source_skus = {str(row.get("sku") or "").strip() for row in prompt_rows if row.get("sku")}
    source_skus.discard("")
    precheck_failed = [row for row in prompt_rows if row.get("precheck_status") == "failed"]

    model_by_sku = {
        str(row.get("sku") or "").strip(): row
        for row in model_rows
        if row.get("sku")
    }
    valid_model_skus: set[str] = set()
    invalid_model_rows: list[tuple[str, str]] = []
    for sku, row in model_by_sku.items():
        is_valid, reason = model_success_is_valid(row)
        if is_valid:
            valid_model_skus.add(sku)
        else:
            invalid_model_rows.append((sku, reason or "模型结果未成功"))

    model_pending_skus = source_skus - set(model_by_sku)
    model_failed_skus = source_skus - valid_model_skus - model_pending_skus

    expected_image_count = len(valid_model_skus) * IMAGE_COUNT
    all_image_input_keys = {
        image_row_key(row)
        for row in image_input_rows
        if row.get("SKU") and row.get("图片命名")
    }
    expected_image_keys = {
        f"{sku}::new_sub{index}_{sku}"
        for sku in valid_model_skus
        for index in range(1, IMAGE_COUNT + 1)
    }

    image_by_key = {
        row_key(row): row
        for row in image_rows
        if row_key(row) in expected_image_keys
    }
    image_status_counts = Counter(str(row.get("status") or "unknown") for row in image_by_key.values())
    image_done_keys = {key for key, row in image_by_key.items() if row.get("status") in {"success", "failed", "missing_file"}}
    image_pending = len(expected_image_keys - image_done_keys)

    upload_by_key = {
        row_key(row): row
        for row in upload_rows
        if row_key(row) in image_by_key and image_by_key[row_key(row)].get("status") == "success"
    }
    upload_expected_keys = {
        key for key, row in image_by_key.items() if row.get("status") == "success"
    }
    upload_status_counts = Counter(str(row.get("status") or "unknown") for row in upload_by_key.values())
    upload_done_keys = {key for key, row in upload_by_key.items() if row.get("status") in {"success", "skipped", "failed", "missing_file"}}
    upload_pending = len(upload_expected_keys - upload_done_keys)
    final_summary = (
        preview_final_image_result_from_logs(paths["oss_results"], paths["image_results"], paths["final_image_excel"])
        if paths["oss_results"].exists()
        else None
    )

    return {
        "batch_name": effective_batch,
        "batch_root": str(paths["root"]),
        "sku_total": len(source_skus),
        "theoretical_image_total": len(source_skus) * IMAGE_COUNT,
        "precheck_failed_skus": len(precheck_failed),
        "precheck_failed_images": len(precheck_failed) * IMAGE_COUNT,
        "model_valid_skus": len(valid_model_skus),
        "model_pending_skus": len(model_pending_skus),
        "model_failed_skus": len(model_failed_skus),
        "model_failed_images": len(model_failed_skus) * IMAGE_COUNT,
        "model_invalid_examples": invalid_model_rows[:10],
        "expected_image_count": expected_image_count,
        "image_input_rows": len(image_input_rows),
        "image_input_duplicate_count": len(image_input_rows) - len(all_image_input_keys),
        "image_input_extra_count": len(all_image_input_keys - expected_image_keys),
        "image_input_missing_count": len(expected_image_keys - all_image_input_keys),
        "image_tracked_count": len(image_by_key),
        "image_status_counts": dict(image_status_counts),
        "image_pending_count": image_pending,
        "upload_expected_count": len(upload_expected_keys),
        "upload_tracked_count": len(upload_by_key),
        "upload_status_counts": dict(upload_status_counts),
        "upload_pending_count": upload_pending,
        "final_result_path": str(paths["final_image_excel"]),
        "final_result_sku_count": final_summary.sku_count if final_summary else 0,
        "final_result_complete_sku_count": final_summary.complete_sku_count if final_summary else 0,
    }


def print_batch_stats(batch_name: str | None = None) -> None:
    stats = collect_stats(batch_name)
    image_counts = stats["image_status_counts"]
    upload_counts = stats["upload_status_counts"]

    print("\n=== 批次统计 ===")
    print(f"批次名: {stats['batch_name']}")
    print(f"批次目录: {stats['batch_root']}")
    print(
        "SKU: "
        f"总数={stats['sku_total']} | "
        f"02有效成功={stats['model_valid_skus']} | "
        f"02待处理={stats['model_pending_skus']} | "
        f"02失败/无效={stats['model_failed_skus']}"
    )
    print(
        "图片: "
        f"理论应生成={stats['theoretical_image_total']} | "
        f"当前可生成={stats['expected_image_count']} | "
        f"03入参行={stats['image_input_rows']} | "
        f"03入参重复={stats['image_input_duplicate_count']} | "
        f"03入参多余={stats['image_input_extra_count']} | "
        f"03入参缺失={stats['image_input_missing_count']}"
    )
    print(
        "03生成下载: "
        f"成功={image_counts.get('success', 0)} | "
        f"已提交={image_counts.get('submitted', 0)} | "
        f"失败={image_counts.get('failed', 0)} | "
        f"待处理={stats['image_pending_count']}"
    )
    print(
        "05上传OSS: "
        f"应上传={stats['upload_expected_count']} | "
        f"成功={upload_counts.get('success', 0)} | "
        f"跳过={upload_counts.get('skipped', 0)} | "
        f"失败={upload_counts.get('failed', 0)} | "
        f"待上传={stats['upload_pending_count']}"
    )
    print(
        "06最终结果: "
        f"SKU={stats['final_result_sku_count']} | "
        f"6张完整SKU={stats['final_result_complete_sku_count']}"
    )

    if stats["precheck_failed_skus"]:
        print(
            "01预检失败: "
            f"SKU={stats['precheck_failed_skus']} | 影响图片={stats['precheck_failed_images']}"
        )
    if stats["model_failed_skus"]:
        print(
            "02失败/无效影响: "
            f"SKU={stats['model_failed_skus']} | 图片={stats['model_failed_images']}"
        )
    examples = stats["model_invalid_examples"]
    if examples:
        print("02失败/无效样例:")
        for sku, reason in examples[:5]:
            print(f"  SKU={sku} | 原因={reason}")

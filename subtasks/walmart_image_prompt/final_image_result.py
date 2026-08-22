"""Build final image result workbook from OSS upload results."""

from __future__ import annotations

import re
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


TASK_ROOT = Path(__file__).resolve().parent
TOTAL_CONFIG_PATH = TASK_ROOT / "config.json"

OUTPUT_COLUMNS = [
    "SKU",
    "处理后主图",
    "处理后附图1",
    "处理后附图2",
    "处理后附图3",
    "处理后附图4",
    "处理后附图5",
    "处理后附图6",
]


def load_image_selection() -> dict[str, Any]:
    """从总配置读取 image_selection（与 build/03 共用同一数据源）。"""
    if not TOTAL_CONFIG_PATH.exists():
        return {}
    try:
        data = json.loads(TOTAL_CONFIG_PATH.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}
    selection = data.get("image_selection") or {}
    return selection if isinstance(selection, dict) else {}


def load_desired_count() -> int:
    """目标张数：image_selection.desired_count；未配置时按 6 张全量。"""
    dc = load_image_selection().get("desired_count")
    return int(dc) if isinstance(dc, int) and dc > 0 else 6


def load_image_type_order() -> list[str]:
    return [str(t).strip() for t in (load_image_selection().get("image_type_order") or []) if str(t).strip()]


@dataclass(slots=True)
class FinalImageResultSummary:
    input_path: Path
    output_path: Path
    source_rows: int
    eligible_rows: int
    sku_count: int
    complete_sku_count: int
    desired_count: int = 6
    missing_by_sku: dict[str, list[str]] | None = None


def build_final_image_result(input_path: str | Path, output_path: str | Path) -> FinalImageResultSummary:
    input_path = Path(input_path)
    output_path = Path(output_path)
    desired_count = load_desired_count()
    rows = read_source_rows(input_path)
    grouped: dict[str, dict[str, Any]] = {}
    eligible_rows = 0

    for row in rows:
        sku = row.get("SKU", "").strip()
        image_name = row.get("图片命名", "").strip()
        upload_url = row.get("OSS上传URL", "").strip()
        main_image = row.get("参考图片链接", "").strip()
        download_status = row.get("下载结果", "").strip()
        upload_status = row.get("OSS上传状态", "").strip()
        image_index = get_sub_index(image_name)

        if not sku or not upload_url or image_index is None:
            continue
        if download_status and download_status != "成功":
            continue
        if upload_status and upload_status not in {"成功", "跳过"}:
            continue

        eligible_rows += 1
        item = grouped.setdefault(
            sku,
            {
                "SKU": sku,
                "处理后主图": main_image,
                **{f"处理后附图{index}": "" for index in range(1, 7)},
            },
        )
        if main_image and not item["处理后主图"]:
            item["处理后主图"] = main_image
        target_column = f"处理后附图{image_index}"
        if not item[target_column]:
            item[target_column] = upload_url

    ordered_rows = list(grouped.values())
    write_result(output_path, ordered_rows)
    complete_sku_count = sum(1 for row in ordered_rows if count_filled_images(row) >= desired_count)
    return FinalImageResultSummary(
        input_path=input_path,
        output_path=output_path,
        source_rows=len(rows),
        eligible_rows=eligible_rows,
        sku_count=len(ordered_rows),
        complete_sku_count=complete_sku_count,
        desired_count=desired_count,
    )


def build_final_image_result_from_logs(
    oss_results_path: str | Path,
    image_results_path: str | Path,
    output_path: str | Path,
) -> FinalImageResultSummary:
    oss_results_path = Path(oss_results_path)
    image_results_path = Path(image_results_path)
    output_path = Path(output_path)
    desired_count = load_desired_count()
    oss_rows = read_jsonl(oss_results_path)
    valid_image_keys = {
        row_key(row)
        for row in read_jsonl(image_results_path)
        if row.get("status") == "success" and row.get("sku") and row.get("image_name")
    }
    reference_map = {
        row_key(row): str(row.get("reference_image") or "").strip()
        for row in read_jsonl(image_results_path)
        if row_key(row) in valid_image_keys
    }
    grouped: dict[str, dict[str, Any]] = {}
    eligible_rows = 0
    seen_keys: set[str] = set()

    for row in oss_rows:
        status = str(row.get("status") or "").strip()
        if status not in {"success", "skipped"}:
            continue
        sku = str(row.get("sku") or "").strip()
        image_name = str(row.get("image_name") or "").strip()
        key = f"{sku}::{image_name}"
        if key not in valid_image_keys or key in seen_keys:
            continue
        upload_url = str(row.get("oss_url") or "").strip()
        image_index = get_sub_index(image_name)
        if not sku or not upload_url or image_index is None:
            continue
        seen_keys.add(key)
        eligible_rows += 1
        item = grouped.setdefault(
            sku,
            {
                "SKU": sku,
                "处理后主图": "",
                **{f"处理后附图{index}": "" for index in range(1, 7)},
            },
        )
        reference_image = reference_map.get(f"{sku}::{image_name}", "")
        if reference_image and not item["处理后主图"]:
            item["处理后主图"] = reference_image
        target_column = f"处理后附图{image_index}"
        if not item[target_column]:
            item[target_column] = upload_url

    ordered_rows = list(grouped.values())
    write_result(output_path, ordered_rows)
    complete_sku_count = count_complete_skus(ordered_rows, desired_count)
    return FinalImageResultSummary(
        input_path=oss_results_path,
        output_path=output_path,
        source_rows=len(oss_rows),
        eligible_rows=eligible_rows,
        sku_count=len(ordered_rows),
        complete_sku_count=complete_sku_count,
        desired_count=desired_count,
        missing_by_sku=collect_missing_types(image_results_path),
    )


def preview_final_image_result_from_logs(
    oss_results_path: str | Path,
    image_results_path: str | Path,
    output_path: str | Path,
) -> FinalImageResultSummary:
    oss_results_path = Path(oss_results_path)
    image_results_path = Path(image_results_path)
    output_path = Path(output_path)
    desired_count = load_desired_count()
    if not oss_results_path.exists():
        return FinalImageResultSummary(oss_results_path, output_path, 0, 0, 0, 0, desired_count=desired_count)
    valid_image_keys = {
        row_key(row)
        for row in read_jsonl(image_results_path)
        if row.get("status") == "success" and row.get("sku") and row.get("image_name")
    }
    grouped: dict[str, set[int]] = {}
    eligible_rows = 0
    seen_keys: set[str] = set()
    for row in read_jsonl(oss_results_path):
        status = str(row.get("status") or "").strip()
        if status not in {"success", "skipped"}:
            continue
        sku = str(row.get("sku") or "").strip()
        image_name = str(row.get("image_name") or "")
        key = f"{sku}::{image_name}"
        if key not in valid_image_keys or key in seen_keys:
            continue
        upload_url = str(row.get("oss_url") or "").strip()
        image_index = get_sub_index(image_name)
        if not sku or not upload_url or image_index is None:
            continue
        seen_keys.add(key)
        eligible_rows += 1
        grouped.setdefault(sku, set()).add(image_index)
    complete_sku_count = sum(1 for indexes in grouped.values() if len(indexes) >= desired_count)
    return FinalImageResultSummary(
        input_path=oss_results_path,
        output_path=output_path,
        source_rows=len(read_jsonl(oss_results_path)),
        eligible_rows=eligible_rows,
        sku_count=len(grouped),
        complete_sku_count=complete_sku_count,
        desired_count=desired_count,
        missing_by_sku=collect_missing_types(image_results_path),
    )


def preview_final_image_result(input_path: str | Path, output_path: str | Path) -> FinalImageResultSummary:
    input_path = Path(input_path)
    output_path = Path(output_path)
    desired_count = load_desired_count()
    rows = read_source_rows(input_path) if input_path.exists() else []
    grouped: dict[str, set[int]] = {}
    eligible_rows = 0
    for row in rows:
        sku = row.get("SKU", "").strip()
        upload_url = row.get("OSS上传URL", "").strip()
        image_index = get_sub_index(row.get("图片命名", ""))
        download_status = row.get("下载结果", "").strip()
        upload_status = row.get("OSS上传状态", "").strip()
        if not sku or not upload_url or image_index is None:
            continue
        if download_status and download_status != "成功":
            continue
        if upload_status and upload_status not in {"成功", "跳过"}:
            continue
        eligible_rows += 1
        grouped.setdefault(sku, set()).add(image_index)
    complete_sku_count = sum(1 for indexes in grouped.values() if len(indexes) >= desired_count)
    return FinalImageResultSummary(
        input_path=input_path,
        output_path=output_path,
        source_rows=len(rows),
        eligible_rows=eligible_rows,
        sku_count=len(grouped),
        complete_sku_count=complete_sku_count,
        desired_count=desired_count,
    )


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def row_key(row: dict[str, Any]) -> str:
    return f"{row.get('sku')}::{row.get('image_name')}"


def count_filled_images(row: dict[str, Any]) -> int:
    return sum(1 for index in range(1, 7) if row.get(f"处理后附图{index}"))


def count_complete_skus(rows: list[dict[str, str]], desired_count: int = 6) -> int:
    return sum(1 for row in rows if count_filled_images(row) >= desired_count)


def collect_missing_types(image_results_path: str | Path) -> dict[str, list[str]]:
    """按 SKU 收集缺失的 image_type：目标未凑满时，候选顺序中未成功生成的部分（供人工补图）。"""
    rows = read_jsonl(Path(image_results_path))
    satisfied: dict[str, set[str]] = {}
    success_count: dict[str, int] = {}
    for row in rows:
        sku = str(row.get("sku") or "").strip()
        image_type = str(row.get("image_type") or "").strip()
        if not sku:
            continue
        if row.get("status") == "success" and image_type:
            satisfied.setdefault(sku, set()).add(image_type)
            success_count[sku] = success_count.get(sku, 0) + 1
    order = load_image_type_order()
    desired_count = load_desired_count()
    missing: dict[str, list[str]] = {}
    for sku, done in satisfied.items():
        if success_count.get(sku, 0) >= desired_count:
            continue
        absent = [t for t in order if t not in done]
        if absent:
            missing[sku] = absent
    return missing


def read_source_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise RuntimeError(f"最终结果输入 Excel 不存在: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [
        str(value).strip() if value is not None and str(value).strip() else ""
        for value in header_values
    ]
    required = {"SKU", "参考图片链接", "图片命名", "OSS上传URL"}
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError("最终结果输入缺少列: " + ", ".join(missing))

    rows: list[dict[str, str]] = []
    for values in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        row = {
            name: "" if values[index] is None else str(values[index]).strip()
            for index, name in enumerate(headers)
            if name
        }
        if any(row.values()):
            rows.append(row)
    return rows


def write_result(path: Path, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(OUTPUT_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def get_sub_index(name: str) -> int | None:
    match = re.search(r"(?:new_)?sub([1-6])", str(name), re.IGNORECASE)
    return int(match.group(1)) if match else None
